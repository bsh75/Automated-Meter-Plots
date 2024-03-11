import xlrd
import openpyxl
from openpyxl import Workbook
from datetime import datetime, timedelta
import os
import matplotlib.pyplot as plt
import calendar
from openpyxl.styles import PatternFill

NAMES = {
    "D43-MCC-B1-2": "Level B1 MCC 2"
    , "D44-MCC-B1-1": "Level B1 MCC 1"
    , "D55-DB-B3": "Level B3 DB"
    , "D4-GEN-1-BUS": "Generator 1"
    , "D5-GEN-2-BUS": "Generator 2"
    , "D6-GEN-3-BUS": "Generator 3"
    , "D9-MCC-7.2": "Level 07 MCC 7.2"
    , "D10-LIFT-DB-3": "Lift 03 DB"
    , "D11-CHILLER-20.3": "Chiller 03 DB"
    , "D12-MCC-20.4": "Level 20 MCC 20.4"
    , "D13-MCC-7.3": "Level 07 MCC 7.3"
    , "D14-DB-SP": "Fire Panel DB"
    , "D15-LIFT-DB-2": "Lift 02 DB"
    , "D16-MCC-20.2": "Level 20 MCC 20.2"
    , "D17-HOUSE-DBs": "House DBs"
    , "D18-DB-CP": "Carpark DB"
    , "D19-MCC-20.3": "Level 20 MCC 20.3"
    , "D20-LIFT-DB-1": "Lift 01 DB"
    , "D22-MCC-19.2": "Level 19 MCC 19.2"
    , "D26-MCC-20.1": "Level 20 MCC 20.1"
    , "D27-CHILLER-20.1": "Chiller 01 DB"
    , "D28-MCC-19.1": "Level 19 MCC 19.1"
    , "D29-DB-R": "Retail 01-04 DB"
    , "D30-MCC-7.1": "Level 07 MCC 7.1"
    , "D31-CHILLER-20.2": "Chiller 02 DB"
    , "D41-DB-RB": "Retail 05 DB"
    , "D42-MCC-G1": "Ground Floor MCC 1"
    , "D7-BNZ-UPS-1": "Level 08 UPS 1"
    , "D8-BNZ-UPS-2": "Level 08 UPS 2"
    , "D32-DB-8": "Level 08 DB"
    , "D33-DB-7K": "Level 07 Kitchen DB"
    , "D34-DB-7": "Level 07 DB"
    , "D35-DB-6": "Level 06 DB"
    , "D36-DB-5": "Level 05 DB"
    , "D37-DB-4": "Level 04 DB"
    , "D38-DB-3": "Level 03 DB"
    , "D39-DB-2": "Level 02 DB"
    , "D40-DB-1": "Level 01 DB"
    , "D57-DB-UDB": "Level 06 DB UDB"
    , "D58-DB-UFB": "Level 07 FDF Supplies L1-8UFB"
    , "D59-DB-UPS-L1": "Level 01 UPS"
    , "D60-DB-UDA": "Level 06 DB UDA"
    , "D61-DB-UFA": "Level 07 FDF Supplies L1-8UFA"
    , "D62-DB-UMA": "Level 06 Machine Room UMA"
    , "D63-DB-UPY": "Level 05 Payments"
    , "D64-L7-Dealers": "Level 07 Dealers"
    , "D65-L6-CommsRoom": "Level 06 Comms Room Cooler"
    , "D66-FDFRoom-CoolerA": "Level 07 FDF Room Cooler A"
    , "D67-FDFRoom-CoolerB": "Level 07 FDF Room Cooler B"
    , "D68-DB-UMB": "Level 06 Machine Room UMB"
    , "D23-DELOITTE-UPS-1": "Level 13 UPS 1"
    , "D24-DELOITTE-UPS-2": "Level 13 UPS 2"
    , "D25-L18-KITCHEN": "Level 18 Kitchen"
    , "D45-DB-18": "Level 18 DB"
    , "D46-DB-17": "Level 17 DB"
    , "D47-DB-16": "Level 16 DB"
    , "D48-DB-15": "Level 15 DB"
    , "D49-DB-14": "Level 14 DB"
    , "D50-DB-13": "Level 13 DB"
    , "D51-DB-12": "Level 12 DB"
    , "D52-DB-11": "Level 11 DB"
    , "D53-DB-10": "Level 10 DB"
    , "D54-DB-9": "Level 09 DB"
}

class Meter:
    '''Defines a class for each meter which contains all the information/data relating to a single meter'''
    def __init__(self, name, dates, off_peaks, on_peaks, weekends, totals):
        self.name = name
        self.dates = [dates]
        self.off_peaks = [off_peaks]
        self.on_peaks = [on_peaks]
        self.weekends = [weekends]
        self.totals = [totals]

    def __str__(self):
        # Return a string representation of the class instance
        return f"MyClass: Name={self.name}, Date={self.dates}, Off Peak={self.off_peaks}, On Peak={self.on_peaks}, Weekend={self.weekends}, Total={self.totals}"
    
class groupedMeters:
    '''Defines a class for a group of meters which contains all the information/data relating to a single meter'''
    def __init__(self, name, dates, off_peaks, on_peaks, weekends, totals):
        self.name = name
        self.dates = dates
        self.off_peaks = off_peaks
        self.on_peaks = on_peaks
        self.weekends = weekends
        self.totals = totals

    
def extract_date_times_from_string(input_string):
    '''Extracts the date from the cell associated with it as a string'''
    # Define the format of the date and time in the string
    date_format = "%B %d, %Y %I:%M %p"

    # Extract the dates and times from the input string
    date_time_str = input_string.split('From:', 1)[1]

    # Split the extracted string to separate the start and end date_time_str
    start_date_time_str, end_date_time_str = date_time_str.split(' to:', 1)

    # Parse the date_time strings into datetime objects
    start_date_time = datetime.strptime(start_date_time_str, date_format)
    end_date_time = datetime.strptime(end_date_time_str.strip(), date_format)

    # Check if the period is exactly one day
    one_day_difference = (end_date_time - start_date_time) == timedelta(days=1)

    if one_day_difference:
        return start_date_time
    else:
        return None

def add_to_meter(name, date, off_p_usage, on_p_usage, weeknd_usage, total_usage, excel_filename, meters_list):
    """Adds the information on meter {name} to that meters class"""
    for meter in meters_list:
        if meter.name == name:
            meter.dates.append(date)
            meter.off_peaks.append(off_p_usage)
            meter.on_peaks.append(on_p_usage)
            meter.weekends.append(weeknd_usage)
            meter.totals.append(total_usage)

def extract_meters_into_class(excel_filename, meters_class_list):
    name_criteria_string = 'Energy User:'  # Replace with the criteria you are looking for
    date_criteria_string = 'For Electric Usage From:'
    off_p_string = 'Off-Peak'
    on_p_string = 'On-Peak'
    weeknd_string = 'Weekend'
    '''Looks in a single file (corresponding to a sinlge date) and creates or adds to existing meter class for each meter found'''
    try:
        # Need to use xlrd for xls (old excel) files
        workbook = xlrd.open_workbook(excel_filename)
        sheet = workbook.sheet_by_index(0)  # Assuming you want to work with the first sheet
    except xlrd.biffh.XLRDError:
        print(f"{excel_filename} is not a xls file, SKIPPING... ")
        return
    # Iterate through all cells to find relevant information
    for row in range(sheet.nrows):
        for col in range(sheet.ncols):
            cell_value = str(sheet.cell_value(row, col))
            # Get the name of meter
            if cell_value.startswith(name_criteria_string):
                name = cell_value.replace(name_criteria_string, '')
            # Get the start of the period as the date for usage
            elif cell_value.startswith(date_criteria_string):
                date = extract_date_times_from_string(cell_value)
                # target_datetime = datetime(2023, 9, 26, 0, 0, 0)
                if not date:
                    print(f"WARNING!! {name} period is not 1 day, SKIPPING file...... ({excel_filename})")
                    return
                # if date == target_datetime:
                #     date = datetime(2023, 8, 26, 0, 0, 0)
                # print(date)
            # Get off peak, on peak, and weekend breakdown
            elif cell_value == off_p_string:
                off_p_usage = sheet.cell_value(row, col+2)
            elif cell_value == on_p_string:
                on_p_usage = sheet.cell_value(row, col+2)
            elif cell_value == weeknd_string:
                # Section should only occur once the weekend value (last bit of info) has been found
                weeknd_usage = sheet.cell_value(row, col+2)
                total_usage = off_p_usage+on_p_usage+weeknd_usage
                if not any(meter.name == name for meter in meters_class_list):
                    meters_class_list.append(Meter(name, date, off_p_usage, on_p_usage, weeknd_usage, total_usage))
                else:
                    add_to_meter(name, date, off_p_usage, on_p_usage, weeknd_usage, total_usage, excel_filename, meters_class_list)

def extract_PN_meters_into_dict(excel_filename, meters_class_list):
    name_criteria_string = 'Energy User:'  # Replace with the criteria you are looking for
    date_criteria_string = 'For Electric Usage From:'
    off_p_string = 'Off-Peak'
    on_p_string = 'On-Peak'
    weeknd_string = 'Weekend'
    '''Looks in a single file (corresponding to a sinlge date) and creates or adds to existing meter class for each meter found'''
    try:
        # Need to use xlrd for xls (old excel) files
        workbook = xlrd.open_workbook(excel_filename)
        sheet = workbook.sheet_by_index(0)  # Assuming you want to work with the first sheet
    except xlrd.biffh.XLRDError:
        print(f"{excel_filename} is not a xls file, SKIPPING... ")
        return
    # Iterate through all cells to find relevant information
    for row in range(sheet.nrows):
        for col in range(sheet.ncols):
            cell_value = str(sheet.cell_value(row, col))
            # Get the name of meter
            if cell_value.startswith(name_criteria_string):
                name = cell_value.replace(name_criteria_string, '')
            # Get the start of the period as the date for usage
            elif cell_value.startswith(date_criteria_string):
                date = extract_date_times_from_string(cell_value)
                # target_datetime = datetime(2023, 9, 26, 0, 0, 0)
                if not date:
                    print(f"WARNING!! {name} period is not 1 day, SKIPPING file...... ({excel_filename})")
                    return
                # if date == target_datetime:
                #     date = datetime(2023, 8, 26, 0, 0, 0)
                # print(date)
            # Get off peak, on peak, and weekend breakdown
            elif cell_value == off_p_string:
                off_p_usage = sheet.cell_value(row, col+2)
            elif cell_value == on_p_string:
                on_p_usage = sheet.cell_value(row, col+2)
            elif cell_value == weeknd_string:
                # Section should only occur once the weekend value (last bit of info) has been found
                weeknd_usage = sheet.cell_value(row, col+2)
                total_usage = off_p_usage+on_p_usage+weeknd_usage
                if not any(meter.name == name for meter in meters_class_list):
                    meters_class_list.append(Meter(name, date, off_p_usage, on_p_usage, weeknd_usage, total_usage))
                else:
                    add_to_meter(name, date, off_p_usage, on_p_usage, weeknd_usage, total_usage, excel_filename, meters_class_list)

def plot_dates_vs_total_from_CLASS(obj, output_filename):
    '''Creates the plots from data associated with a single class'''
    # Get the dates and totals from the object
    dates = obj.dates
    off_ps = obj.off_peaks
    on_ps = obj.on_peaks
    wknds = obj.weekends
    totals = obj.totals
    
    # Create the plot
    fig, ax1 = plt.subplots(figsize=(10, 6))

    # Create a new y-axis for plotting the accumulation data
    ax2 = ax1.twinx()

    # Set constant limits to the Y axis of each
    if 'Basement' in obj.name:
        ax1.set_ylim(0, 500)
        ax2.set_ylim(0, 12200)
    elif 'Common' in obj.name:
        ax1.set_ylim(0, 20000)
        ax2.set_ylim(0, 400000)
    elif 'Level 01' in obj.name:
        ax1.set_ylim(0, 3700)
        ax2.set_ylim(0, 80000)
    elif 'Level 09' in obj.name:
        ax1.set_ylim(0, 2500)
        ax2.set_ylim(0, 50000)

    # # Convert the dates to datetime objects
    dates_strings = [datetime.strftime(date, "%d-%b-%y") for date in dates]

    # Plot the usage data on the left y-axis
    p3 = ax1.bar(dates, wknds, color='#6794a7', width=0.6, label='Weekend')
    p1 = ax1.bar(dates, off_ps, bottom=wknds, color='#7ad2f6', width=0.6, label='Off Peak')
    p2 = ax1.bar(dates, on_ps, bottom=[sum(x) for x in zip(wknds, off_ps)], color='#014d64', width=0.6, label='On Peak')

    ax1.set_xlabel('Day')
    ax1.set_ylabel('Usage (kwH)')
    # ax1.set_title(f'Dates vs. Usage for {obj.name}')
    # ax1.grid(axis='y', linestyle='--', alpha=0.7)


    # Calculate the accumulation data (cumulative sum of totals)
    accumulations = [sum(totals[:i + 1]) for i in range(len(totals))]

    # Plot the accumulation data on the right y-axis
    ax2.plot(dates, accumulations, color='#90353B', marker='o', label='Accumulation')
    ax2.set_ylabel('Accumulation (kwH)')

    # Combine the handles for bars and lines for a single legend
    handles, labels = ax1.get_legend_handles_labels()
    handles2, labels2 = ax2.get_legend_handles_labels()
    ax1.legend(handles + handles2, labels + labels2, loc='upper left')

    # Rotate x-axis labels for better visibility
    ax1.set_xticks(dates)
    # ax1.set_xticklabels(dates_strings, rotation=45, ha='right')
    ax1.set_xticklabels([i for i in range(1, len(dates)+1)])

    # Save the plot
    plt.tight_layout()
    output_filename = output_filename.replace('.', '_')
    print(f"Saving image file with name: {output_filename}")
    plt.savefig(output_filename + '.png')
    # plt.show()
    plt.close()

    # Save the data used in the plots
    wb = Workbook()
    sheet = wb.active
    sheet.append([f"Data for meter: {obj.name}"])
    dates_strings.insert(0, "Dates")
    off_ps.insert(0, "Off Peak")
    on_ps.insert(0, "On Peak")
    wknds.insert(0, "Weekend")
    totals.insert(0, "Total")
    sheet.append(dates_strings)
    sheet.append(off_ps)
    sheet.append(on_ps)
    sheet.append(wknds)
    sheet.append(totals)
    # for i in range(0, len(dates)):
    #     sheet.append([dates_strings[i], off_ps[i], on_ps[i], wknds[i], totals[i]])
    wb.save(f"{output_filename}.xlsx")

def set_uniform_spacing(worksheet, start_column, end_column, width):
    """Sets all the column widths between start column and end column to be width"""
    for col_idx in range(start_column, end_column + 1):
        column_letter = openpyxl.utils.get_column_letter(col_idx)
        worksheet.column_dimensions[column_letter].width = width

def update_merged_cell_value(worksheet, row, column, value):
    """Updates the merged cell containing (row, col) with the value"""
    # Get the merged cell range containing the cell at (row, column)
    for cell_range in worksheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row  = cell_range.bounds
        if (min_row <= row <= max_row) and (min_col <= column <= max_col):
            # Unmerge the cells within the range
            worksheet.unmerge_cells(str(cell_range))
            # Update the value in each cell within the merged cell range
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    worksheet.cell(row=r, column=c, value=value)
            # Merge the cells again
            worksheet.merge_cells(start_row=min_row, start_column=min_col,
                                  end_row=max_row, end_column=max_col)
            # Break the loop after updating the merged cell range
            break

def is_weekend_day(date_string):
    """Returns true if date_string is in a weekend"""
    # Convert the date string to a date object
    date_obj = datetime.strptime(date_string, "%d-%b-%y").date()
    # Check if the day of the week is Saturday (5) or Sunday (6)
    return date_obj.weekday() in [5, 6]

def write_all_data_to_excel(meters_class_list, file_path, month, output_sheet_name):
    """Writes the data for all meters into a copy of the template sheet"""
    wb = openpyxl.load_workbook(file_path)
    
    # Check if a sheet with the same name already exists
    if output_sheet_name in wb.sheetnames:
        wb.remove(wb[output_sheet_name])  # Remove the existing sheet
    
    template_sheet = wb["template"]

    # Create a new sheet thats a copy of the template one
    new_sheet_name = output_sheet_name
    sheet = wb.copy_worksheet(template_sheet)
    sheet.title = new_sheet_name

    # Updates the merged cell containing the month
    update_merged_cell_value(sheet, 1, 3, month)

    # Get the number of days in the given month
    month_num_days = calendar.monthrange(2023, list(calendar.month_abbr).index(month[:3]))[1]
    print(month_num_days)

    # Delete excess columns in the template sheet
    heading_cols = 2 # 1 = A, 2 = B, 3 = C
    month_col_start = heading_cols + 1
    month_col_end = sheet.max_column -1 # Minus the total column
    print(month_col_end)
    if (month_col_end - heading_cols) > month_num_days:
        delete_cols = sheet.iter_cols(min_col=heading_cols + month_num_days, max_col=month_col_end)
        for col in delete_cols:
            month_col_end -= 1
            sheet.delete_cols(col[0].column)

    wb.save(file_path)

    # Get the dates in the first row and format them in the same format as meter dates
    dates_nums = [sheet.cell(row=2, column=col_idx).value for col_idx in range(month_col_start, month_col_end+1)]
    # print(dates_nums)
    month_year = meters_class_list[0].dates[0].strftime("%d-%b-%y")[-6:]
    dates_string = [f'{str(date_num).zfill(2)}-{month_year}' for date_num in dates_nums]
    # print(dates_string)
    # Create a dictionary to map between the two date formats
    date_map = {}
    for idx, date_s in enumerate(dates_string):
        date_map[date_s] = idx + month_col_start  # Match the column index
    # print(date_map)
    weekend_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')

    # Find the column indices of the weekend dates
    weekend_cols = [col_idx for col_idx, date in enumerate(dates_string, start=month_col_start) if is_weekend_day(date)]

    # Write data to the table
    row_idx = 3 # Data entry starts at row 3 (after month and days rows)
    for meter in meters_class_list:
        # Write the meter name in the first column (is a merged column)
        meter_name = meter.name
        update_merged_cell_value(sheet, row_idx, 1, meter_name)

        for i in range(0, len(meter.dates)):
            date_str = meter.dates[i].strftime("%d-%b-%y")
            off_p_val = meter.off_peaks[i]
            on_p_val = meter.on_peaks[i]
            wknd_val = meter.weekends[i]
            total_val = meter.totals[i]
            col_idx = date_map[date_str]

            cell = sheet.cell(row=row_idx, column=col_idx, value=off_p_val)
            cell = sheet.cell(row=row_idx+1, column=col_idx, value=on_p_val)
            cell = sheet.cell(row=row_idx+2, column=col_idx, value=wknd_val)
            cell = sheet.cell(row=row_idx+3, column=col_idx, value=total_val)

        # Apply the weekend_fill to the entire row for the weekend dates
        for col in weekend_cols:
            weekend_cell = sheet.cell(row=row_idx, column=col)
            weekend_cell.fill = weekend_fill
            weekend_cell1 = sheet.cell(row=row_idx+1, column=col)
            weekend_cell1.fill = weekend_fill
            weekend_cell2 = sheet.cell(row=row_idx+2, column=col)
            weekend_cell2.fill = weekend_fill
            weekend_cell3 = sheet.cell(row=row_idx+3, column=col)
            weekend_cell3.fill = weekend_fill

        row_idx += 4 # As 4 rows have been filled

    # Sets unifor spacing for the range specified
    set_uniform_spacing(sheet, month_col_start, month_col_end + 2, width=5)

    wb.save(file_path)

def write_all_data_grouped_to_excel(meters_class_dict_byGroup, file_path, month, output_sheet_name, dates_row):
    """Writes the data for all meters into a copy of the template sheet"""
    wb = openpyxl.load_workbook(file_path)
    
    # Check if a sheet with the same name already exists
    if output_sheet_name in wb.sheetnames:
        wb.remove(wb[output_sheet_name])  # Remove the existing sheet
    
    template_sheet = wb["Power Meters - ALL - TEMPLATE"]

    # Create a new sheet thats a copy of the template one
    new_sheet_name = output_sheet_name
    sheet = wb.copy_worksheet(template_sheet)
    sheet.title = new_sheet_name

    # Updates the merged cell containing the month
    update_merged_cell_value(sheet, 3, 3, month)

    # Get the number of days in the given month
    # month_num_days = calendar.monthrange(2023, list(calendar.month_abbr).index(month[:3]))[1]
    month_num_days = len(meters_class_dict_byGroup[next(iter(meters_class_dict_byGroup))][0].dates)
    
    # Delete excess columns in the template sheet
    heading_cols = 2 # 1 = A, 2 = B, 3 = C
    month_col_start = heading_cols + 1
    last_month_col = 31 + heading_cols # Minus the total column
    if (last_month_col - heading_cols) > month_num_days:
        for i in range(month_col_start+month_num_days, last_month_col+1):
            print(sheet.cell(row=dates_row, column=i))
            cell = sheet.cell(row=dates_row, column=i, value='NA')
        # delete_cols = sheet.iter_cols(min_col=month_col_start + month_num_days, max_col=last_month_col)
        # for col in delete_cols:
        #     last_month_col -= 1
        #     # sheet.delete_cols(col[0].column)

    wb.save(file_path)

    col_end = month_num_days + month_col_start

    # Get the dates in the first row and format them in the same format as meter dates
    dates_nums = [sheet.cell(row=dates_row, column=col_idx).value for col_idx in range(month_col_start, col_end)]
    month_year = meters_class_dict_byGroup['Basement'][0].dates[0].strftime("%d-%b-%y")[-6:]
    dates_string = [f'{str(date_num).zfill(2)}-{month_year}' for date_num in dates_nums]
    # Create a dictionary to map between the two date formats
    date_map = {}
    for idx, date_s in enumerate(dates_string):
        date_map[date_s] = idx + month_col_start  # Match the column index
    # print(date_map)
    weekend_fill = PatternFill(start_color='D3DEF1', end_color='D3DEF1', fill_type='solid')

    # Find the column indices of the weekend dates
    weekend_cols = [col_idx for col_idx, date in enumerate(dates_string, start=month_col_start) if is_weekend_day(date)]

    # Write data to the table
    row_idx = dates_row +1 # Data entry starts at row 3 (after month and days rows)
    for group_name, meter_list in meters_class_dict_byGroup.items():
        # Udate the 
        update_merged_cell_value(sheet, row_idx, 1, group_name)
        row_idx += 1
        for meter in meter_list:

            # Write the meter name in the first column (is a merged column)
            meter_name = NAMES[meter.name]
            update_merged_cell_value(sheet, row_idx, 1, meter_name)
            for i in range(0, len(meter.dates)):
                date_str = meter.dates[i].strftime("%d-%b-%y")
                off_p_val = meter.off_peaks[i]
                on_p_val = meter.on_peaks[i]
                wknd_val = meter.weekends[i]
                total_val = meter.totals[i]
                col_idx = date_map[date_str]

                cell = sheet.cell(row=row_idx, column=col_idx, value=off_p_val)
                cell = sheet.cell(row=row_idx+1, column=col_idx, value=on_p_val)
                cell = sheet.cell(row=row_idx+2, column=col_idx, value=wknd_val)
                cell = sheet.cell(row=row_idx+3, column=col_idx, value=total_val)

            # Apply the weekend_fill to the entire row for the weekend dates
            for col in weekend_cols:
                weekend_cell = sheet.cell(row=row_idx, column=col)
                weekend_cell.fill = weekend_fill
                weekend_cell1 = sheet.cell(row=row_idx+1, column=col)
                weekend_cell1.fill = weekend_fill
                weekend_cell2 = sheet.cell(row=row_idx+2, column=col)
                weekend_cell2.fill = weekend_fill
                weekend_cell3 = sheet.cell(row=row_idx+3, column=col)
                weekend_cell3.fill = weekend_fill

            row_idx += 4 # As 4 rows have been filled

    # Sets uniform spacing for the range specified
    set_uniform_spacing(sheet, month_col_start, col_end + 2, width=5)

    wb.save(file_path)



def write_groups_to_excel(grouped_meter_list, file_path, month, output_sheet_name, dates_row):
    """Writes the data for all meters into a copy of the template sheet"""
    wb = openpyxl.load_workbook(file_path)
    
    # Check if a sheet with the same name already exists
    if output_sheet_name in wb.sheetnames:
        wb.remove(wb[output_sheet_name])  # Remove the existing sheet
    
    template_sheet = wb["Power Meters - Group - TEMPLATE"]

    # Create a new sheet thats a copy of the template one
    new_sheet_name = output_sheet_name
    sheet = wb.copy_worksheet(template_sheet)
    sheet.title = new_sheet_name

    # Updates the merged cell containing the month
    update_merged_cell_value(sheet, 3, 3, month)

    # Get the number of days in the given month
    # month_num_days = calendar.monthrange(2023, list(calendar.month_abbr).index(month[:3]))[1]
    month_num_days = len(grouped_meter_list[0].dates)
    
    # Delete excess columns in the template sheet
    heading_cols = 2 # 1 = A, 2 = B, 3 = C
    month_col_start = heading_cols + 1
    last_month_col = 31 + heading_cols # Minus the total column
    if (last_month_col - heading_cols) > month_num_days:
        for i in range(month_col_start+month_num_days, last_month_col+1):
            print(sheet.cell(row=dates_row, column=i))
            cell = sheet.cell(row=dates_row, column=i, value='NA')
        # delete_cols = sheet.iter_cols(min_col=month_col_start + month_num_days, max_col=last_month_col)
        # for col in delete_cols:
        #     last_month_col -= 1
        #     # sheet.delete_cols(col[0].column)

    wb.save(file_path)

    col_end = month_num_days + month_col_start

    # Get the dates in the first row and format them in the same format as meter dates
    dates_nums = [sheet.cell(row=dates_row, column=col_idx).value for col_idx in range(month_col_start, col_end)]
    month_year = grouped_meter_list[0].dates[0].strftime("%d-%b-%y")[-6:]
    dates_string = [f'{str(date_num).zfill(2)}-{month_year}' for date_num in dates_nums]
    # Create a dictionary to map between the two date formats
    date_map = {}
    for idx, date_s in enumerate(dates_string):
        date_map[date_s] = idx + month_col_start  # Match the column index
    # print(date_map)
    weekend_fill = PatternFill(start_color='D3DEF1', end_color='D3DEF1', fill_type='solid')

    # Find the column indices of the weekend dates
    weekend_cols = [col_idx for col_idx, date in enumerate(dates_string, start=month_col_start) if is_weekend_day(date)]
    # Write data to the table
    row_idx = dates_row + 1 # Start index of data entry

    for meter in grouped_meter_list:
        # print(meter.name)
        update_merged_cell_value(sheet, row_idx, 1, meter.name)
        row_idx += 1
        # print(meter.on_peaks)
        for i in range(0, len(meter.dates)):
            date_str = meter.dates[i].strftime("%d-%b-%y")
            off_p_val = meter.off_peaks[i]
            on_p_val = meter.on_peaks[i]
            wknd_val = meter.weekends[i]
            total_val = meter.totals[i]
            col_idx = date_map[date_str]
            
            cell = sheet.cell(row=row_idx, column=col_idx, value=off_p_val)
            cell = sheet.cell(row=row_idx+1, column=col_idx, value=on_p_val)
            cell = sheet.cell(row=row_idx+2, column=col_idx, value=wknd_val)
            cell = sheet.cell(row=row_idx+3, column=col_idx, value=total_val)

        # Apply the weekend_fill to the entire row for the weekend dates
        for col in weekend_cols:
            weekend_cell = sheet.cell(row=row_idx, column=col)
            weekend_cell.fill = weekend_fill
            weekend_cell1 = sheet.cell(row=row_idx+1, column=col)
            weekend_cell1.fill = weekend_fill
            weekend_cell2 = sheet.cell(row=row_idx+2, column=col)
            weekend_cell2.fill = weekend_fill
            weekend_cell3 = sheet.cell(row=row_idx+3, column=col)
            weekend_cell3.fill = weekend_fill

        row_idx += 4 # As 4 rows have been filled

    # Sets uniform spacing for the range specified
    # set_uniform_spacing(sheet, month_col_start, col_end + 2, width=5)

    wb.save(file_path)


def find_group_name(className, meters_name_dict):
    for groupName, substringList in meters_name_dict.items():
        for substring in substringList:
            # print(substring)
            # print(className)
            if (substring + '-') in className:
                return groupName

def split_meter_list_into_groups(meters_dict, meterC_list):
    meter_class_grouped_dict = {
        'Basement': [],
        'Common Areas': [],            
        'Level 01-08': [],
        'Level 09-18': []
    }
    for meterC in meterC_list:
        class_name = meterC.name
        group_name = find_group_name(class_name, meters_dict)
        if group_name:
            meter_class_grouped_dict[group_name].append(meterC)
    return meter_class_grouped_dict