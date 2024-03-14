import xlrd
import openpyxl
from openpyxl import Workbook
from datetime import datetime, timedelta
import os
import matplotlib.pyplot as plt
import calendar
from openpyxl.styles import PatternFill
import pandas as pd
# from functions import *


class Meter:
    '''Defines a class for each meter which contains all the information/data relating to a single meter'''
    def __init__(self, name, dates, off_peaks, on_peaks, weekends, totals):
        self.name = name
        self.dates = dates
        self.off_peaks = off_peaks
        self.on_peaks = on_peaks
        self.weekends = weekends
        self.totals = totals

    def __str__(self):
        # Return a string representation of the class instance
        return f"MyClass: Name={self.name}, Date={self.dates}, Off Peak={self.off_peaks}, On Peak={self.on_peaks}, Weekend={self.weekends}, Total={self.totals}"

NAMES = {
    "D43-MCC-B1-2": "Level B1 MCC 2"
    , "D44-MCC-B1-1": "Level B1 MCC 1"
    , "D55-DB-B3": "Level B3 DB"
    , "D4-GEN-1-BUS": "Generator 1"
    , "D5-GEN-2-BUS": "Generator 2"
    , "D6-GEN-3-BUS": "Generator 3"
    , "D9-MCC-7.2": "Level 07 MCC 7.2"
    , "D10-LIFT-DB-3": "Lift 03 DB"
    , "D11-CHILLER-20.3": "Chiller 03 DB"
    , "D12-MCC-20.4": "Level 20 MCC 20.4"
    , "D13-MCC-7.3": "Level 07 MCC 7.3"
    , "D14-DB-SP": "Fire Panel DB"
    , "D15-LIFT-DB-2": "Lift 02 DB"
    , "D16-MCC-20.2": "Level 20 MCC 20.2"
    , "D17-HOUSE-DBs": "House DBs"
    , "D18-DB-CP": "Carpark DB"
    , "D19-MCC-20.3": "Level 20 MCC 20.3"
    , "D20-LIFT-DB-1": "Lift 01 DB"
    , "D22-MCC-19.2": "Level 19 MCC 19.2"
    , "D26-MCC-20.1": "Level 20 MCC 20.1"
    , "D27-CHILLER-20.1": "Chiller 01 DB"
    , "D28-MCC-19.1": "Level 19 MCC 19.1"
    , "D29-DB-R": "Retail 01-04 DB"
    , "D30-MCC-7.1": "Level 07 MCC 7.1"
    , "D31-CHILLER-20.2": "Chiller 02 DB"
    , "D41-DB-RB": "Retail 05 DB"
    , "D42-MCC-G1": "Ground Floor MCC 1"
    , "D7-BNZ-UPS-1": "Level 08 UPS 1"
    , "D8-BNZ-UPS-2": "Level 08 UPS 2"
    , "D32-DB-8": "Level 08 DB"
    , "D33-DB-7K": "Level 07 Kitchen DB"
    , "D34-DB-7": "Level 07 DB"
    , "D35-DB-6": "Level 06 DB"
    , "D36-DB-5": "Level 05 DB"
    , "D37-DB-4": "Level 04 DB"
    , "D38-DB-3": "Level 03 DB"
    , "D39-DB-2": "Level 02 DB"
    , "D40-DB-1": "Level 01 DB"
    , "D57-DB-UDB": "Level 06 DB UDB"
    , "D58-DB-UFB": "Level 07 FDF Supplies L1-8UFB"
    , "D59-DB-UPS-L1": "Level 01 UPS"
    , "D60-DB-UDA": "Level 06 DB UDA"
    , "D61-DB-UFA": "Level 07 FDF Supplies L1-8UFA"
    , "D62-DB-UMA": "Level 06 Machine Room UMA"
    , "D63-DB-UPY": "Level 05 Payments"
    , "D64-L7-Dealers": "Level 07 Dealers"
    , "D65-L6-CommsRoom": "Level 06 Comms Room Cooler"
    , "D66-FDFRoom-CoolerA": "Level 07 FDF Room Cooler A"
    , "D67-FDFRoom-CoolerB": "Level 07 FDF Room Cooler B"
    , "D68-DB-UMB": "Level 06 Machine Room UMB"
    , "D23-DELOITTE-UPS-1": "Level 13 UPS 1"
    , "D24-DELOITTE-UPS-2": "Level 13 UPS 2"
    , "D25-L18-KITCHEN": "Level 18 Kitchen"
    , "D45-DB-18": "Level 18 DB"
    , "D46-DB-17": "Level 17 DB"
    , "D47-DB-16": "Level 16 DB"
    , "D48-DB-15": "Level 15 DB"
    , "D49-DB-14": "Level 14 DB"
    , "D50-DB-13": "Level 13 DB"
    , "D51-DB-12": "Level 12 DB"
    , "D52-DB-11": "Level 11 DB"
    , "D53-DB-10": "Level 10 DB"
    , "D54-DB-9": "Level 09 DB"
}

def combine_meters_in_list_to_class(grouped_METER_list, grouped_meter_name):
    base_meter = grouped_METER_list[0]
    dates = base_meter.dates
    lol_off_peaks = [base_meter.off_peaks]
    lol_on_peaks = [base_meter.on_peaks]
    lol_weekends = [base_meter.weekends]
    lol_totals = [base_meter.totals]
    
    print(f'\n\n\n{lol_off_peaks}\n\n\n')
    for meter in grouped_METER_list[1:]:
        lol_off_peaks.append(meter.off_peaks)
        lol_on_peaks.append(meter.on_peaks)
        lol_weekends.append(meter.weekends)
        lol_totals.append(meter.totals)

    # print(lol_off_peaks)
    total_off_peaks = [sum(x) for x in zip(*lol_off_peaks)]
    total_on_peaks = [sum(x) for x in zip(*lol_on_peaks)]
    total_weekends = [sum(x) for x in zip(*lol_weekends)]
    total_totals = [sum(x) for x in zip(*lol_totals)]
        
    return Meter(grouped_meter_name, dates, total_off_peaks, total_on_peaks, total_weekends, total_totals)


def find_group_name(className, meters_name_dict):
    for groupName, substringList in meters_name_dict.items():
        for substring in substringList:
            if (substring + '-') in className:
                return groupName
            

def split_meter_list_into_groups(meterC_list):
    meter_groups_dict = {
    'Basement': ['D44', 'D43', 'D55'],
    'Common Areas': ['D4', 'D5', 'D6', 'D27', 'D31', 'D11', 'D20', 'D15', 'D10', 'D14', 'D17', 'D18', 'D29', 'D41', 'D42', 'D30', 'D9', 'D13', 'D28', 'D22', 'D26', 'D16', 'D19', 'D12'],            
    'Level 01-08': ['D40', 'D59', 'D39', 'D38', 'D37', 'D36', 'D63', 'D35', 'D60', 'D57', 'D62', 'D68', 'D65', 'D34', 'D33', 'D61', 'D58', 'D66', 'D67', 'D64', 'D32', 'D7', 'D8'],
    'Level 09-18': [
        'D54','D53','D52','D51','D50','D23','D24','D49','D48','D47','D46','D45','D25']
    }
    meter_class_grouped_dict = {
        'Basement': [],
        'Common Areas': [],            
        'Level 01-08': [],
        'Level 09-18': []
    }
    for meterC in meterC_list:
        class_name = meterC.name
        group_name = find_group_name(class_name, meter_groups_dict)
        if group_name:
            meter_class_grouped_dict[group_name].append(meterC)
    return meter_class_grouped_dict


def plot_dates_vs_total_from_CLASS(obj, output_filename, save_plot_data_to_sheet):
    '''Creates the plots from data associated with a single class'''
    # Get the dates and totals from the object
    date_strings = obj.dates
    off_ps = obj.off_peaks
    on_ps = obj.on_peaks
    wknds = obj.weekends
    totals = obj.totals
    
    # Create the plot
    fig, ax1 = plt.subplots(figsize=(10, 6))

    # Create a new y-axis for plotting the accumulation data
    ax2 = ax1.twinx()

    # Set constant limits to the Y axis of each
    if 'Basement' in obj.name:
        ax1.set_ylim(0, 500)
        ax2.set_ylim(0, 12200)
    elif 'Common' in obj.name:
        ax1.set_ylim(0, 20000)
        ax2.set_ylim(0, 400000)
    elif 'Level 01' in obj.name:
        ax1.set_ylim(0, 3700)
        ax2.set_ylim(0, 80000)
    elif 'Level 09' in obj.name:
        ax1.set_ylim(0, 2500)
        ax2.set_ylim(0, 50000)

    # # Convert the dates to datetime objects
    dates = [datetime.strptime(date, "%d-%b-%y") for date in date_strings]

    # Plot the usage data on the left y-axis
    p3 = ax1.bar(dates, wknds, color='#6794a7', width=0.6, label='Weekend')
    p1 = ax1.bar(dates, off_ps, bottom=wknds, color='#7ad2f6', width=0.6, label='Off Peak')
    p2 = ax1.bar(dates, on_ps, bottom=[sum(x) for x in zip(wknds, off_ps)], color='#014d64', width=0.6, label='On Peak')

    ax1.set_xlabel('Day')
    ax1.set_ylabel('Usage (kwH)')

    # Calculate the accumulation data (cumulative sum of totals)
    accumulations = [sum(totals[:i + 1]) for i in range(len(totals))]

    # Plot the accumulation data on the right y-axis
    ax2.plot(dates, accumulations, color='#90353B', marker='o', label='Accumulation')
    ax2.set_ylabel('Accumulation (kwH)')

    # Combine the handles for bars and lines for a single legend
    handles, labels = ax1.get_legend_handles_labels()
    handles2, labels2 = ax2.get_legend_handles_labels()
    ax1.legend(handles + handles2, labels + labels2, loc='upper left')

    # Rotate x-axis labels for better visibility
    ax1.set_xticks(dates)
    # ax1.set_xticklabels(dates_strings, rotation=45, ha='right')
    ax1.set_xticklabels([i for i in range(1, len(dates)+1)])

    # Save the plot
    plt.tight_layout()
    output_filename = output_filename.replace('.', '_')
    print(f"Saving image file with name: {output_filename}")
    plt.savefig(output_filename + '.png')
    # plt.show()
    plt.close()

    # Save the data used in the plots
    if save_plot_data_to_sheet:
        wb = Workbook()
        sheet = wb.active
        sheet.append([f"Data for meter: {obj.name}"])
        date_strings.insert(0, "Dates")
        off_ps.insert(0, "Off Peak")
        on_ps.insert(0, "On Peak")
        wknds.insert(0, "Weekend")
        totals.insert(0, "Total")
        sheet.append(date_strings)
        sheet.append(off_ps)
        sheet.append(on_ps)
        sheet.append(wknds)
        sheet.append(totals)
        # for i in range(0, len(dates)):
        #     sheet.append([dates_strings[i], off_ps[i], on_ps[i], wknds[i], totals[i]])
        wb.save(f"{output_filename}.xlsx")


def update_merged_cell_value(worksheet, row, column, value):
    """Updates the merged cell containing (row, col) with the value"""
    # Get the merged cell range containing the cell at (row, column)
    for cell_range in worksheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row  = cell_range.bounds
        if (min_row <= row <= max_row) and (min_col <= column <= max_col):
            # Unmerge the cells within the range
            worksheet.unmerge_cells(str(cell_range))
            # Update the value in each cell within the merged cell range
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    worksheet.cell(row=r, column=c, value=value)
            # Merge the cells again
            worksheet.merge_cells(start_row=min_row, start_column=min_col,
                                  end_row=max_row, end_column=max_col)
            # Break the loop after updating the merged cell range
            break


def is_weekend_day(date_string):
    """Returns true if date_string is in a weekend"""
    # Convert the date string to a date object
    date_obj = datetime.strptime(date_string, "%d-%b-%y").date()
    # Check if the day of the week is Saturday (5) or Sunday (6)
    return date_obj.weekday() in [5, 6]


def write_groups_to_excel(grouped_meter_list, file_path, month, output_sheet_name, dates_row):
    """Writes the data for all meters into a copy of the template sheet"""
    wb = openpyxl.load_workbook(file_path)
    
    # Check if a sheet with the same name already exists
    if output_sheet_name in wb.sheetnames:
        wb.remove(wb[output_sheet_name])  # Remove the existing sheet
    
    template_sheet = wb["Power Meters - Group - TEMPLATE"]

    # Create a new sheet thats a copy of the template one
    new_sheet_name = output_sheet_name
    sheet = wb.copy_worksheet(template_sheet)
    sheet.title = new_sheet_name

    # Updates the merged cell containing the month
    update_merged_cell_value(sheet, 3, 3, month)

    # Get the number of days in the given month
    # month_num_days = calendar.monthrange(2023, list(calendar.month_abbr).index(month[:3]))[1]
    month_num_days = len(grouped_meter_list[0].dates)
    
    # Delete excess columns in the template sheet
    heading_cols = 2 # 1 = A, 2 = B, 3 = C
    month_col_start = heading_cols + 1
    last_month_col = 31 + heading_cols # Minus the total column
    if (last_month_col - heading_cols) > month_num_days:
        for i in range(month_col_start+month_num_days, last_month_col+1):
            print(sheet.cell(row=dates_row, column=i))
            cell = sheet.cell(row=dates_row, column=i, value='NA')
        # delete_cols = sheet.iter_cols(min_col=month_col_start + month_num_days, max_col=last_month_col)
        # for col in delete_cols:
        #     last_month_col -= 1
        #     # sheet.delete_cols(col[0].column)

    wb.save(file_path)

    col_end = month_num_days + month_col_start

    # Get the dates in the first row and format them in the same format as meter dates
    dates_nums = [sheet.cell(row=dates_row, column=col_idx).value for col_idx in range(month_col_start, col_end)]
    month_year = grouped_meter_list[0].dates[0][-6:]
    dates_string = [f'{str(date_num).zfill(2)}-{month_year}' for date_num in dates_nums]
    # Create a dictionary to map between the two date formats
    date_map = {}
    for idx, date_s in enumerate(dates_string):
        # print(date_s, '  ', date_map)
        date_map[date_s] = idx + month_col_start  # Match the column index
    # print(date_map)
    weekend_fill = PatternFill(start_color='D3DEF1', end_color='D3DEF1', fill_type='solid')

    # Find the column indices of the weekend dates
    weekend_cols = [col_idx for col_idx, date in enumerate(dates_string, start=month_col_start) if is_weekend_day(date)]
    # Write data to the table
    row_idx = dates_row + 1 # Start index of data entry

    for meter in grouped_meter_list:
        # print(meter.name)
        update_merged_cell_value(sheet, row_idx, 1, meter.name)
        row_idx += 1
        # print(meter.on_peaks)
        for i in range(0, len(meter.dates)):
            date_str = meter.dates[i]
            off_p_val = meter.off_peaks[i]
            on_p_val = meter.on_peaks[i]
            wknd_val = meter.weekends[i]
            total_val = meter.totals[i]
            col_idx = date_map[date_str]
            
            cell = sheet.cell(row=row_idx, column=col_idx, value=off_p_val)
            cell = sheet.cell(row=row_idx+1, column=col_idx, value=on_p_val)
            cell = sheet.cell(row=row_idx+2, column=col_idx, value=wknd_val)
            cell = sheet.cell(row=row_idx+3, column=col_idx, value=total_val)

        # Apply the weekend_fill to the entire row for the weekend dates
        for col in weekend_cols:
            weekend_cell = sheet.cell(row=row_idx, column=col)
            weekend_cell.fill = weekend_fill
            weekend_cell1 = sheet.cell(row=row_idx+1, column=col)
            weekend_cell1.fill = weekend_fill
            weekend_cell2 = sheet.cell(row=row_idx+2, column=col)
            weekend_cell2.fill = weekend_fill
            weekend_cell3 = sheet.cell(row=row_idx+3, column=col)
            weekend_cell3.fill = weekend_fill

        row_idx += 4 # As 4 rows have been filled

    # Sets uniform spacing for the range specified
    # set_uniform_spacing(sheet, month_col_start, col_end + 2, width=5)

    wb.save(file_path)


def write_all_data_grouped_to_excel(meters_class_dict_byGroup, file_path, month, output_sheet_name, dates_row):
    """Writes the data for all meters into a copy of the template sheet"""
    wb = openpyxl.load_workbook(file_path)
    
    # Check if a sheet with the same name already exists
    if output_sheet_name in wb.sheetnames:
        wb.remove(wb[output_sheet_name])  # Remove the existing sheet
    
    template_sheet = wb["Power Meters - ALL - TEMPLATE"]

    # Create a new sheet thats a copy of the template one
    new_sheet_name = output_sheet_name
    sheet = wb.copy_worksheet(template_sheet)
    sheet.title = new_sheet_name

    # Updates the merged cell containing the month
    update_merged_cell_value(sheet, 3, 3, month)

    # Get the number of days in the given month
    # month_num_days = calendar.monthrange(2023, list(calendar.month_abbr).index(month[:3]))[1]
    month_num_days = len(meters_class_dict_byGroup[next(iter(meters_class_dict_byGroup))][0].dates)
    
    # Delete excess columns in the template sheet
    heading_cols = 2 # 1 = A, 2 = B, 3 = C
    month_col_start = heading_cols + 1
    last_month_col = 31 + heading_cols # Minus the total column
    if (last_month_col - heading_cols) > month_num_days:
        for i in range(month_col_start+month_num_days, last_month_col+1):
            print(sheet.cell(row=dates_row, column=i))
            cell = sheet.cell(row=dates_row, column=i, value='NA')
        # delete_cols = sheet.iter_cols(min_col=month_col_start + month_num_days, max_col=last_month_col)
        # for col in delete_cols:
        #     last_month_col -= 1
        #     # sheet.delete_cols(col[0].column)

    wb.save(file_path)

    col_end = month_num_days + month_col_start

    # Get the dates in the first row and format them in the same format as meter dates
    dates_nums = [sheet.cell(row=dates_row, column=col_idx).value for col_idx in range(month_col_start, col_end)]
    month_year = meters_class_dict_byGroup['Basement'][0].dates[0][-6:]
    dates_string = [f'{str(date_num).zfill(2)}-{month_year}' for date_num in dates_nums]
    # Create a dictionary to map between the two date formats
    date_map = {}
    for idx, date_s in enumerate(dates_string):
        date_map[date_s] = idx + month_col_start  # Match the column index
    # print(date_map)
    weekend_fill = PatternFill(start_color='D3DEF1', end_color='D3DEF1', fill_type='solid')

    # Find the column indices of the weekend dates
    weekend_cols = [col_idx for col_idx, date in enumerate(dates_string, start=month_col_start) if is_weekend_day(date)]

    # Write data to the table
    row_idx = dates_row +1 # Data entry starts at row 3 (after month and days rows)
    for group_name, meter_list in meters_class_dict_byGroup.items():
        # Udate the 
        update_merged_cell_value(sheet, row_idx, 1, group_name)
        row_idx += 1
        for meter in meter_list:

            # Write the meter name in the first column (is a merged column)
            meter_name = NAMES[meter.name]
            update_merged_cell_value(sheet, row_idx, 1, meter_name)
            for i in range(0, len(meter.dates)):
                date_str = meter.dates[i]
                off_p_val = meter.off_peaks[i]
                on_p_val = meter.on_peaks[i]
                wknd_val = meter.weekends[i]
                total_val = meter.totals[i]
                col_idx = date_map[date_str]

                cell = sheet.cell(row=row_idx, column=col_idx, value=off_p_val)
                cell = sheet.cell(row=row_idx+1, column=col_idx, value=on_p_val)
                cell = sheet.cell(row=row_idx+2, column=col_idx, value=wknd_val)
                cell = sheet.cell(row=row_idx+3, column=col_idx, value=total_val)

            # Apply the weekend_fill to the entire row for the weekend dates
            for col in weekend_cols:
                weekend_cell = sheet.cell(row=row_idx, column=col)
                weekend_cell.fill = weekend_fill
                weekend_cell1 = sheet.cell(row=row_idx+1, column=col)
                weekend_cell1.fill = weekend_fill
                weekend_cell2 = sheet.cell(row=row_idx+2, column=col)
                weekend_cell2.fill = weekend_fill
                weekend_cell3 = sheet.cell(row=row_idx+3, column=col)
                weekend_cell3.fill = weekend_fill

            row_idx += 4 # As 4 rows have been filled

    # # Sets uniform spacing for the range specified
    # set_uniform_spacing(sheet, month_col_start, col_end + 2, width=5)

    wb.save(file_path)



### INPUTS
month_folder_input = input("Enter the name of the month folder eg. '2023 11 November': ")
month_folder = f'POWER/{month_folder_input}'

current_month = month_folder_input.split(' ')[-1]
cur_month_input = input(f"If the month you would like to process is not {current_month}, type the month now... otherwise press 'Enter'")
if cur_month_input != '':
    current_month = cur_month_input

meter_data_folder = input("Is the name of the data folder 'each_meter_data' if not, enter it here: ")

if meter_data_folder == '':
    meter_data_folder = 'each_meter_data'

input_data_path = f"{month_folder}/{meter_data_folder}"
print(f"Processing {input_data_path} ...")

save_plot_data = False
save_input = input("Would you like to save the plot data into separate xlsx file (Y/N): ")
if save_input.lower() == 'y':
    save_plot_data = True

### OUTPUTS
output_data_folder = 'Outputs'
output_data_path = f"{month_folder}/{output_data_folder}"
if not os.path.exists(output_data_path):
    os.makedirs(output_data_path)

all_METERS_list = []
### Get a class for each meter to store the data
input_files = os.listdir(input_data_path)
for filename in input_files:
    if filename.endswith(".xlsx"):
        file_path = f"{input_data_path}/{filename}"
        df = pd.read_excel(file_path)
        name = df.columns[0]
        dates = df.iloc[0].tolist()[1:]
        off_p_usages = df.iloc[1].tolist()[1:]
        on_p_usages = df.iloc[2].tolist()[1:]
        weeknd_usages = df.iloc[3].tolist()[1:]
        total_usages = df.iloc[4].tolist()[1:]

        meter_class = Meter(name, dates, off_p_usages, on_p_usages, weeknd_usages, total_usages)
        print(meter_class)
        all_METERS_list.append(meter_class)



group_classes_to_plot = []
grouped_meterslist_dict = split_meter_list_into_groups(all_METERS_list)
print(grouped_meterslist_dict)
for key, value in grouped_meterslist_dict.items():
    group_meter = combine_meters_in_list_to_class(grouped_METER_list=value, grouped_meter_name=key)
    # print('\n', group_meter.name, group_meter.on_peaks)
    """BELOW IS MAIN FUNCITON FOR CREATING PLOTS"""
    plot_dates_vs_total_from_CLASS(group_meter, output_filename=f"{output_data_path}/{group_meter.name}", save_plot_data_to_sheet=save_plot_data)
    group_classes_to_plot.append(combine_meters_in_list_to_class(grouped_METER_list=value, grouped_meter_name=key))
    print(group_meter, '\n', value, ' ---- ', key)

"""BELOW IS MAIN FUNCITON FOR WRITING DATA TO EXCEL"""
write_groups_to_excel(group_classes_to_plot, 'POWER/80 Queen St - Analytics Report - Calendar Data.xlsx', current_month, current_month+'-grouped', dates_row=3)
write_all_data_grouped_to_excel(grouped_meterslist_dict, 'POWER/80 Queen St - Analytics Report - Calendar Data.xlsx', current_month, current_month, dates_row=4)
""""""