import csv
import openpyxl
from datetime import datetime
import os

def get_meters_data(ws, gap):
    """Function which looks at a worksheet to find its name and data"""
    name_start = 'history:BMS_Supervisor/YDA_'
    start_indexs = []
    location_names = []
    name_data_dict = {}
    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if str(row[0]).startswith(name_start):
            start_indexs.append(row_idx)
            location_names.append(str(row[0]).replace(name_start, ''))

    if len(start_indexs) > 1:
        for i in range(0, len(start_indexs)-1):
            name_data_dict[location_names[i]] = [list(row) for row in ws.iter_rows(min_row=start_indexs[i] + gap + 1,
                                                                                  max_row=start_indexs[i + 1],
                                                                                  values_only=True,
                                                                                  min_col=1,
                                                                                  max_col=4)]
    data_to_add = [list(row) for row in ws.iter_rows(min_row=start_indexs[-1] + gap + 1,
                                                     values_only=True,
                                                     min_col=1,
                                                     max_col=4)]
    name_data_dict[location_names[-1]] = data_to_add
    return name_data_dict

def trim_data_dictionary(dictionary, month):
    """ Takes a dictionary containing the meters and all the timepoints for them and returns a similar dictionary 
        containing only the useful timepoints (12am 11:55pm and the first and last)"""
    new_dictionary = {}
    for meter, data_set in dictionary.items():
        # Get the dates and values 
        dates = [date.replace(' NZST', '') for date, _, _, _ in data_set]
        values = [int(value[0:-3]) for _, _, _, value in data_set]

        # Limit the dates to be just the month requested
        for i in range(0, len(dates)):
            if month in dates[i]:
                break
        dates = dates[i:]
        values = values[i:]

        wanted_dates_values = []
        
        # The first date and value are set as the past date/value
        past_date = dates[0][:9]
        past_value = values[0]

        for i in range(1, len(dates)):
            current_date = dates[i][:9]
            current_value = values[i]
            if (current_date != past_date):
                # If current date is different from the past, then take the difference in values to be the past date's usage
                wanted_dates_values.append([past_date, past_value])
                past_date = current_date
                past_value = current_value
            else:
                continue

        # Add the last entry (and also get the last time for the last date)
        wanted_dates_values.append([dates[-1][:9], values[-1]])
        new_dictionary[meter] = wanted_dates_values

    return new_dictionary

start_sequence = '80QAfterhoursYesterday'
file_type = '.xlsx'
folder = '80Q - Afterhours Usage'
output_folder = 'Plot Data'
name_gap = 4
file_list = os.listdir(folder)
desired_month = 'Jul'

for file in file_list:
    if file.endswith(file_type):
        print(f"working on file: {file}")
    if file.endswith(file_type) and file.startswith(start_sequence):
        input_xlsx_path = f'{folder}/{file}'
        wb = openpyxl.load_workbook(input_xlsx_path)
        ws = wb.active
        mins_data_dict = get_meters_data(ws, name_gap)
        date_data_dict = trim_data_dictionary(mins_data_dict, desired_month)
        print(date_data_dict)
