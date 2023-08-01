import os
from openpyxl import load_workbook


def transpose_xlsx_files(folder_path):
    # Get a list of all .xlsx files in the folder
    xlsx_files = [file for file in os.listdir(folder_path) if file.endswith('.xlsx')]

    for file in xlsx_files:
        file_path = os.path.join(folder_path, file)
        # Load the workbook
        wb = load_workbook(file_path)
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            transposed_data = list(zip(*sheet.iter_rows(values_only=True)))
            
            # Create a new sheet with the name "{original sheet name} - transposed"
            new_sheet_name = f"{sheet_name} - transposed"
            new_sheet = wb.create_sheet(title=new_sheet_name)
            
            # Write the transposed data to the new sheet
            for row in transposed_data:
                new_sheet.append(row)
        
        # Save the modified workbook
        wb.save(file_path)

# Example usage:
folder_path = "80Q - Power Meters - Powernet Data\Plot Data"
transpose_xlsx_files(folder_path)
