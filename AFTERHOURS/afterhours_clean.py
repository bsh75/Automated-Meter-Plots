# Step 1: Read the CSV file and extract columns A and D
import csv
import pandas as pd
import matplotlib.pyplot as plt
from datetime import datetime
import os
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule, ColorScale
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import calendar

print(openpyxl.__version__)


def get_ahrs_data(df, gap):
    """Function which looks at a file to find its the names and data (suits for multiple items in one file)"""
    name_start = 'history:BMS_Supervisor/YDA_'
    start_indexs = []
    location_names = []
    # data_sets = []
    name_data_dict = {}
    for index, row in df.iterrows():
        if str(row[0]).startswith(name_start):
            start_indexs.append(index)
            location_names.append(df.iloc[index,0].replace(name_start, ''))

    # Now go through each of the found meters and add the data to the dictionary for each one
    if len(start_indexs) > 1:
        for i in range(0, len(start_indexs)-1):
            name_data_dict[location_names[i]] = df.iloc[start_indexs[i]+gap:start_indexs[i+1], [0, 3]]
    data_to_add = df.iloc[start_indexs[-1]+gap:, [0, 3]]
    name_data_dict[location_names[-1]] = data_to_add
    print(name_data_dict)
    return name_data_dict

    # with open(filename, 'r') as csvfile:
    #     csv_reader = csv.reader(csvfile)
    #     for row in csv_reader:
    #         if len(row) >= header_index:
    #             columns_A_and_D.append([row[0], row[3]])

    # return columns_A_and_D

def get_relevant_info(time_usage_cols):
    """Returns all the dates and usages for 12AM and 11:55pm"""
    date_val_pair = []
    # Index for adding the first timepoint regardless of its time (Skips the header)
    i = 0
    for row in time_usage_cols[1:]:
        if ('11:55:00 PM' in row[0]) or ('12:00:00 AM' in row[0]) or (i == 0):
            i += 1
            date = row[0][:9] # Always in DD-MMM-YY format
            value = int(row[1][:-4]) # Removes trailing units and converts to usable number
            date_val_pair.append([date, value])

    if len(date_val_pair) % 2 != 0:
        print("WARNING -- Non even entries")

    return date_val_pair

def plot_water_usage_with_accumulation(data, meter, output_filename):
    dates = [datetime.strptime(date_str, "%d-%b-%y") for date_str, _ in data]
    water_usage = [usage for _, usage in data]

    # Create the primary axes for the bar graph
    fig, ax1 = plt.subplots(figsize=(10, 6))

    # Plot the bar graph on the primary axes
    ax1.bar(dates, water_usage, align='center', alpha=0.7, label='Water Usage')
    ax1.set_xlabel('Date')
    ax1.set_ylabel('Water Usage ($m^3$)')  # Using LaTeX notation for m^3
    ax1.set_title(f'{meter} Water Usage')
    ax1.set_xticks(dates)
    ax1.set_xticklabels([date.strftime("%d-%b-%y") for date in dates], rotation=45, ha='right')

    # Create the secondary axes for the accumulation line
    ax2 = ax1.twinx()

    # Accumulate the water usage data
    accumulated_usage = [sum(water_usage[:i + 1]) for i in range(len(water_usage))]

    # Plot the accumulation line on the secondary axes
    ax2.plot(dates, accumulated_usage, color='red', label='Accumulation')
    ax2.set_ylabel('Accumulation ($m^3$)')
    
    # Combine the legend for both axes
    lines, labels = ax1.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    ax2.legend(lines + lines2, labels + labels2, loc='upper left')

    plt.tight_layout()
    plt.savefig(output_filename)
    plt.close()

def trim_data_dictionary(dictionary, month):
    """ Takes a dictionary containing the floors and all the timepoints for them and returns a similar dictionary 
        containing only the useful timepoints (12am 11:55pm and the first and last)"""
    new_dictionary = {}
    for floor, data_set in dictionary.items():
        # Get the dates and values 
        dates = [date.replace(' NZDT', '') for date in data_set['dates_col'].tolist()] # Remove the " NZDT"
        values = [int(value[0:-3]) for value in data_set['values_col'].tolist()]

        wanted_dates_values = []
        first = True

        for i in range(1, len(dates)):
            current_date = dates[i][:9]
            if month in current_date:
                if first:
                    wanted_dates_values.append((current_date, values[i+1])) # Add the first (date, value) Tup
                    past_date = dates[i][:9] # Gets the date only (not time) 
                    first = False
                    month_yr_str = current_date[3:]
                elif (current_date != past_date):
                    # If current date is different to the past the take the difference in values to be the past dates usage
                    wanted_dates_values.append((current_date, values[i+1]))
                    past_date = current_date
                else:
                    continue
            else:
                continue
        new_dictionary[floor] = wanted_dates_values
    print('\n',new_dictionary,'\n\n')
    return new_dictionary, month_yr_str

def set_uniform_spacing(worksheet, start_column, end_column, width):
    for col_idx in range(start_column, end_column + 1):
        column_letter = openpyxl.utils.get_column_letter(col_idx)
        worksheet.column_dimensions[column_letter].width = width

def update_merged_cell_value(worksheet, row, column, value):
    # Get the merged cell range containing the cell at (row, column)
    for cell_range in worksheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row  = cell_range.bounds
        if (min_row <= row <= max_row) and (min_col <= column <= max_col):
            # Unmerge the cells within the range
            worksheet.unmerge_cells(str(cell_range))
            # Update the value in each cell within the merged cell range
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    worksheet.cell(row=r, column=c, value=value)
            # Merge the cells again
            worksheet.merge_cells(start_row=min_row, start_column=min_col,
                                  end_row=max_row, end_column=max_col)
            # Break the loop after updating the merged cell range
            break

def is_weekend_day(date_string):
    # Convert the date string to a date object
    date_obj = datetime.strptime(date_string, "%d-%b-%y").date()

    # Check if the day of the week is Saturday (5) or Sunday (6)
    return date_obj.weekday() in [5, 6]

def write_data_to_excel(dictionary, file_path, month):
    wb = openpyxl.load_workbook(file_path)
    template_sheet = wb["template"]
    
    new_sheet_name = month
    sheet = wb.copy_worksheet(template_sheet)
    sheet.title = new_sheet_name

    update_merged_cell_value(sheet, 1, 2, month)

    # Get the number of days in the given month
    month_num_days = calendar.monthrange(2023, list(calendar.month_abbr).index(month[:3]))[1]
    # month_num_days = 28

    # Delete excess columns in the template sheet
    col_start = 2
    col_end = sheet.max_column -1
    if col_end > month_num_days + 1:
        # Create a list of columns to delete in reverse order
        cols_to_delete = [col_idx for col_idx in range(col_start + month_num_days + 1, col_end + 1)]
        for col_idx in cols_to_delete:
            print(f"Deleting col: {col_idx-1}")
            sheet.delete_cols(col_idx-1)

    wb.save(file_path)

    col_start = 2
    col_end = sheet.max_column - 1

    # Get the dates in the first row starting from column B and format them in the same format as dictionary dates
    dates = [sheet.cell(row=2, column=col_idx).value for col_idx in range(col_start, col_end)]
    month_year = next(iter(dictionary.values()))[0][0][3:]
    dates = [f'{str(date).zfill(2)}-{month_year}' for date in dates]

    # Create a dictionary to map between the two date formats
    date_map = {}
    for idx, date in enumerate(dates):
        num_date = date
        date_map[num_date] = idx + 2  # Add 2 to match the column index

    weekend_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')

    # Find the column indices of the weekend dates
    weekend_cols = [col_idx for col_idx, date in enumerate(dates, start=col_start) if is_weekend_day(date)]

    # Write data to the table
    row_idx = 3
    for floor, values in dictionary.items():
        # Write the values corresponding to the dates in the table
        for date_str, value in values:
            if desired_month in date_str:
                col_idx = date_map[date_str]
                # Remove any 0's
                if value == 0:
                    value = ''
                cell = sheet.cell(row=row_idx, column=col_idx, value=value)

        # Apply the weekend_fill to the entire row for the weekend dates
        for col in weekend_cols:
            weekend_cell = sheet.cell(row=row_idx, column=col)
            weekend_cell.fill = weekend_fill

        row_idx += 1

    set_uniform_spacing(sheet, col_start, col_end + 2, 4)

    wb.save(file_path)

start_sequence = 'Level'
file_type = '.csv'
folder = input("Enter the name of the folder containing raw data: ")
month_guess = folder[-3:]
month_input = input(f"Is '{month_guess}' the correct month? (Type 'Y' or abbreviated month): ")
if month_input == "Y":
    desired_month = month_guess
else:
    desired_month = month_input

# folder = 'AFTERHOURS/Aug-Sep'AFTERHOURS/Aug-Sep

name_gap = 4
file_list = os.listdir(folder)

floor_dict_List = []

for file in file_list:
    if file.endswith(file_type) and file.startswith(start_sequence):
        # If file is suitable then find a dictionary containing all the different items in that file, and their date_date tuples
        input_csv_path = f'{folder}/{file}'
        data_frame = pd.read_csv(input_csv_path, names=["dates_col", "col2", "col3", "values_col"])
        floor_data_dict = get_ahrs_data(data_frame, name_gap)
        date_usage_dict, month_year = trim_data_dictionary(floor_data_dict, desired_month)
        floor_dict_List.append(date_usage_dict)
        print('\n', date_usage_dict)

# Combine data to be just one dictionary for ease of use
single_dictionary = {}
for floor_dict in floor_dict_List:
    for meter, date_data in floor_dict.items():
        single_dictionary[meter] = date_data

template_xlsx = 'after_hours_tables_local.xlsx'

write_data_to_excel(single_dictionary, template_xlsx, month_year)