import pandas as pd
from datetime import datetime
import os
import openpyxl
import calendar
from openpyxl.styles import Font, PatternFill
import matplotlib.pyplot as plt
from datetime import datetime, timedelta
import re

def remove_leading_zeros(meter_string):
    # Use regular expression to match 'M' followed by one or more digits and remove leading zeros
    return re.sub(r'M0+(\d+)', r'M\1', meter_string)


def process_csv_files(folder_path, desired_month):
    file_list = os.listdir(folder_path)
    meter_data_dict = {}
    start_index = 4
    for file_name in file_list:
        # Read the CSV file into a pandas DataFrame
        file_path = f'{folder_path}/{file_name}'
        df = pd.read_csv(file_path, names=['A', 'B', 'C', 'D'])
        # print(df)

        # Get the date and time from A3
        date_time_ampm_zone = df.iloc[2, 0].split()
        date_str = date_time_ampm_zone[0]
        time_str = date_time_ampm_zone[1] + date_time_ampm_zone[2]
        # print(date_str)

        if (desired_month not in date_str) or ('PreviousMonth' in file_name):
            print(desired_month, date_str)
            continue

        # Convert date and time strings to datetime objects
        date_time_str = f"{date_str} {time_str}"
        date_time = datetime.strptime(date_time_str, '%d-%b-%y %I:%M%p')

        # Get the total usage from B3, C3, and D3
        start_values = df.loc[start_index:, 'B']
        end_values = df.loc[start_index:, 'C']
        total_usages = df.loc[start_index:, 'D'].tolist()
        # Get the meters from col A starting from row 5
        meters = df.loc[start_index:, 'A'].tolist()
        # print(meters)
        # Add the date/total usage pair to each meter in the dictionary
        for i in range(0, len(meters)):
            meter = remove_leading_zeros(meters[i])
            total_usage = float(total_usages[i])
            if meter in meter_data_dict:
                meter_data_dict[meter].append((date_str, total_usage))
                # print("should be adding to meter")
            else:
                meter_data_dict[meter] = [(date_str, total_usage)]
    print(meter_data_dict)
    return meter_data_dict


def get_actual_name(meter_code):
    name_map_dict = {
    'BNZ Floors': 'M1',
    'Deloitte':'M2',
    'Cooling Towers':'M3',
    'BNZ retail':'M4',
    'Altezano Café':'M5',
    'Lacoste':'M6',
    'Ben Sherman':'M7',
    'Rockport':'M8',
    'North Face':'M9',
    'Loading dock':'M10',
    'BNZ Showers':'M11',
    'Deloitte Showers':'M12',
    'Harvest BNZ Showers':'M13',
    'Harvest Deloitte Showers': 'M14',
    'Harvest BNZ Retail': 'M15',
    'Harvest B1 & BNZ': 'M16',
    'Harvest - Ground Flr': 'M17',
    'Harvest Deloitte': 'M18',
    'Harvest not used': 'M19',
     'Harvest Domestic top up': 'M20'}
    # Map the meterName from the input data to the actual name if none found then keep code
    meterName = meter_code
    for actual_name, code in name_map_dict.items():
        if meter_code == code:
            meterName = actual_name + f"({meterName})"
    return meterName


def set_uniform_spacing(worksheet, start_column, end_column, width):
    for col_idx in range(start_column, end_column + 1):
        column_letter = openpyxl.utils.get_column_letter(col_idx)
        worksheet.column_dimensions[column_letter].width = width


def update_merged_cell_value(worksheet, row, column, value):
    # Get the merged cell range containing the cell at (row, column)
    for cell_range in worksheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row  = cell_range.bounds
        if (min_row <= row <= max_row) and (min_col <= column <= max_col):
            # Unmerge the cells within the range
            worksheet.unmerge_cells(str(cell_range))
            # Update the value in each cell within the merged cell range
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    worksheet.cell(row=r, column=c, value=value)
            # Merge the cells again
            worksheet.merge_cells(start_row=min_row, start_column=min_col,
                                  end_row=max_row, end_column=max_col)
            # Break the loop after updating the merged cell range
            break


def is_weekend_day(date_string):
    # Convert the date string to a date object
    date_obj = datetime.strptime(date_string, "%d-%b-%y").date()

    # Check if the day of the week is Saturday (5) or Sunday (6)
    return date_obj.weekday() in [5, 6]


def write_data_to_excel(dictionary, file_path, month, sheet_name):
    '''Writes the dictionary data into the excel sheet containing all the meters (and/or groups)'''
    wb = openpyxl.load_workbook(file_path)
    
    # Check if a sheet with the same name already exists
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])  # Remove the existing sheet
    
    if 'Group' in sheet_name:
        template_sheet = wb["groupingTemplate"]  # Assuming the data should be written to Sheet1
    else:
        template_sheet = wb["template"]  # Assuming the data should be written to Sheet1

    new_sheet_name = sheet_name
    sheet = wb.copy_worksheet(template_sheet)
    sheet.title = new_sheet_name

    update_merged_cell_value(sheet, 1, 2, month)

    # Get the number of days in the given month
    month_num_days = calendar.monthrange(2023, list(calendar.month_abbr).index(month[:3]))[1]

    # Delete excess columns in the template sheet
    heading_cols = 1 # 1 = A, 2 = B, 3 = C
    month_col_start = heading_cols + 1
    month_col_end = sheet.max_column -1 # Minus the total column
    print(month_col_end)
    if (month_col_end - heading_cols) > month_num_days:
        delete_cols = sheet.iter_cols(min_col=heading_cols + month_num_days, max_col=month_col_end)
        for col in delete_cols:
            month_col_end -= 1
            sheet.delete_cols(col[0].column)

    wb.save(file_path)

    # Get the dates in the first row starting from column B and format them in the same format as dictionary dates
    dates = [sheet.cell(row=2, column=col_idx).value for col_idx in range(month_col_start, month_col_end+1)]
    month_year = next(iter(dictionary.values()))[0][0][3:]
    dates = [f'{str(date).zfill(2)}-{month_year}' for date in dates]

    # Create a dictionary to map between the two date formats
    date_map = {}
    for idx, date in enumerate(dates):
        num_date = date
        date_map[num_date] = idx + 2  # Add 2 to match the column index
    print(date_map)
    weekend_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')

    # Find the column indices of the weekend dates
    weekend_cols = [col_idx for col_idx, date in enumerate(dates, start=month_col_start) if is_weekend_day(date)]

    # Write data to the table
    row_idx = 3
    for meter_code, values in dictionary.items():
        # Write the meter name in the first column
        meter_name = get_actual_name(meter_code)
        sheet.cell(row=row_idx, column=1, value=meter_name)

        # Write the values corresponding to the dates in the table
        for date_str, value in values:
            col_idx = date_map[date_str]
            # Remove any 0's
            if value == 0:
                value = ''
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)

        # Apply the weekend_fill to the entire row for the weekend dates
        for col in weekend_cols:
            weekend_cell = sheet.cell(row=row_idx, column=col)
            weekend_cell.fill = weekend_fill

        row_idx += 1

    set_uniform_spacing(sheet, month_col_start, month_col_end + 2, 4)

    wb.save(file_path)


def group_meters(original_dict, groups_dict):
    """
    Group meters based on specified combinations, add and subtract values accordingly.

    Parameters:
        original_dict (dict): Original dictionary containing meter names as keys and date_value pairs as values.
        groups_dict (dict): A dictionary where the keys are the names of the subgroups, and the values are lists
                            containing two lists: the first list contains meters whose values should be added together,
                            and the second list contains meters whose values should be subtracted.

    Returns:
        dict: A new dictionary where each key is a subgroup name, and the corresponding value is a list of
              aggregated date_value pairs for all the meters in that subgroup, considering the specified additions
              and subtractions.
    """
    def aggregate_values(meter_list):
        aggregated_data = {}
        for meter in meter_list:
            if meter in original_dict:
                for date, value in original_dict[meter]:
                    if date in aggregated_data:
                        aggregated_data[date] += value
                    else:
                        aggregated_data[date] = value
        return [(date, value) for date, value in aggregated_data.items()]

    grouped_dict = {}

    for group_name, group_lists in groups_dict.items():
        add_meters = group_lists[0]
        subtract_meters = group_lists[1]

        grouped_data = aggregate_values(add_meters)

        if subtract_meters:
            subtracted_data = aggregate_values(subtract_meters)
            for date, value in subtracted_data:
                for idx, (d, v) in enumerate(grouped_data):
                    if d == date:
                        grouped_data[idx] = (d, v - value)
                        break
                else:
                    print("SOMETHING FIHSY")
                    grouped_data.append((date, -value))

        grouped_dict[group_name] = grouped_data

    return grouped_dict


def plot_water_usage_with_accumulation(data, meter, output_filename):
    dates = [datetime.strptime(date_str, "%d-%b-%y") for date_str, _ in data]
    water_usage = [usage for _, usage in data]

    # Create the primary axes for the bar graph
    fig, ax1 = plt.subplots(figsize=(10, 6))

    # Plot the bar graph on the primary axes
    ax1.bar(dates, water_usage, align='center', alpha=0.7, label='Water Usage')
    ax1.set_xlabel('Date')
    ax1.set_ylabel('Water Usage ($m^3$)')  # Using LaTeX notation for m^3
    ax1.set_title(f'{meter} Water Usage')
    ax1.set_xticks(dates)
    ax1.set_xticklabels([date.strftime("%d-%b-%y") for date in dates], rotation=45, ha='right')

    # Create the secondary axes for the accumulation line
    ax2 = ax1.twinx()

    # Accumulate the water usage data
    accumulated_usage = [sum(water_usage[:i + 1]) for i in range(len(water_usage))]

    # Plot the accumulation line on the secondary axes
    ax2.plot(dates, accumulated_usage, color='red', label='Accumulation')
    ax2.set_ylabel('Accumulation ($m^3$)')
    
    # Combine the legend for both axes
    lines, labels = ax1.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    ax2.legend(lines + lines2, labels + labels2, loc='upper left')

    plt.tight_layout()
    plt.savefig(output_filename)
    plt.close()


def pad_missing_dates(data):
    # Convert the date strings to datetime objects for easy comparison
    dates = [datetime.strptime(row[0], '%d-%b-%y') for row in data[1:-1]]
    datas_dates = [pair[0] for pair in data]
    values = [pair[1] for pair in data]

    # Find the start and end dates in the data
    start_date = min(dates)
    end_date = max(dates)

    # Create a list to store the modified data
    modified_data = [data[0]]  # Header row

    # Loop through the dates and add missing dates with 'NA'
    current_date = start_date
    i = 0
    while current_date <= end_date:
        formatted_date = current_date.strftime('%d-%b-%y')
        if formatted_date not in datas_dates:
            # print(formatted_date, '----', data)
            modified_data.append([formatted_date, 'NA'])
        else:
            modified_data.append([datas_dates[i], values[i]])
            i += 1

        # Move to the next date
        current_date += timedelta(days=1)

    # Add the last row
    modified_data.append(data[-1])

    return modified_data

def plot_each_meter(meters_dictionary):
    first = True
    for meterName, date_usage in meters_dictionary.items():

        meterName = get_actual_name(meterName)

        timestamp_format = "%d-%b-%y %I:%M:%S %p"

        output_xlsx_filename = f'{output_folder_path}/{meterName}.xlsx'
        plot_output_name = f'{output_folder_path}/{meterName}.png'

        print(f'New files created {output_xlsx_filename} and plot {plot_output_name}')
        plot_water_usage_with_accumulation(date_usage, meterName, plot_output_name)
        # date_usage.insert(0, ['Date',  'Water Usage (m\u00b3)'])
        # date_usage.append(['Up until: ', end_time])
        output_df = pd.DataFrame(date_usage)
        output_df.to_excel(output_xlsx_filename, index=False)
        date_usage = pad_missing_dates(date_usage)

        # Add data to the water meter table 
        # Get the list of values
        values = [row[1] for row in date_usage[1:-1]]
        values.insert(0, meterName)
        if first:
            # Get the list of dates
            dates_header_for_table = [row[0] for row in date_usage[:-1]]
            first = False
            water_meter_table_data.append(dates_header_for_table)
        water_meter_table_data.append(values)
    return water_meter_table_data


# Folder pointing to all the water meter data from Niagara
input_folder_path = 'WATER/Aug2023/Water_Data'
output_folder_path = 'WATER/Aug2023/Water_Plot_Data'
desired_month = 'Aug'
# Gets a dictionary containing each meter, and a list of (datetime, usage) pairs for each
meters_data_dict_to_plot = process_csv_files(input_folder_path, desired_month)
water_meter_table_data = []

# water_meter_table = plot_each_meter(meters_data_dict_to_plot)

template_xlsx = f'WATER/all_water_meters_table.xlsx'

write_data_to_excel(meters_data_dict_to_plot, template_xlsx, month=desired_month, sheet_name=desired_month)

subgroup_add_sub_dict = {}
subgroup_add_sub_dict['Total Building (M1+M4:M10)'] = [['M1', 'M10', 'M4', 'M5', 'M6', 'M7', 'M8', 'M9'], []]
subgroup_add_sub_dict['Deloitte (M2+M3+M12)'] = [['M2', 'M3', 'M12'], []]
subgroup_add_sub_dict['BNZ (M1+M2+M4+M11)'] = [['M1', 'M2', 'M11', 'M4'], []]
subgroup_add_sub_dict['True Alliance (M6:M9)'] = [['M6', 'M7', 'M8', 'M9'], []]
subgroup_add_sub_dict['Altezano Cafe (M5)'] = [['M5'], []]
subgroup_add_sub_dict['Basement & Harvest topup (M10-M11-M12)'] = [['M10'], ['M11', 'M12']]
subgroup_add_sub_dict['Recovered Water (M13:M18-M20)'] = [['M13', 'M14', 'M15', 'M16', 'M17', 'M18'], ['M20']]
subgroup_add_sub_dict['HVAC (M3)'] = [['M3'], []]
subgroup_add_sub_dict['BNZ Flushing (M13+M16+M15)'] = [['M13', 'M15', 'M16'], []]
subgroup_add_sub_dict['Deloitte Flushing (M14+M18)'] = [['M14', 'M18'], []]

sub_group_dict = group_meters(meters_data_dict_to_plot, subgroup_add_sub_dict)
for subgroup, data in sub_group_dict.items():
    print(f"{subgroup}\n{data}")

write_data_to_excel(sub_group_dict, template_xlsx, desired_month, f'Grouping for {desired_month}')