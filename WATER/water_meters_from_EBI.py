# Step 1: Read the CSV file and extract columns A and D
import csv
import os
import pandas as pd
import matplotlib.pyplot as plt
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill

def modded_csv_to_xlsx_data_return(csv_filename, month):
    # Read CSV data and convert the second column to integers
    csv_data = []
    with open(csv_filename, 'r', encoding='utf-8-sig') as csv_file:  # Use utf-8-sig to remove BOM
        csv_reader = csv.reader(csv_file)
        first_row = next(csv_reader)  # Read and store the first row
        for row in csv_reader:
            # Convert the second column to an integer
            row[1] = int(row[1])
            if month in row[0]:
                csv_data.append(row)
    return csv_data

def extract_meter_values(dataframe):
    meter_dict = {}
    
    for meter in dataframe.columns[2:]:
        # Extract the meter name by removing the prefix and suffix
        meter_name = meter[len('HYD-METER-'):-len('-TZ.PV')]
        
        date_time_list = []
        value_list = []
        for index, row in dataframe.iterrows():
            date_str = pd.to_datetime(row['Date']).strftime('%d-%b-%y')
            # print(row['Time'])
            # time_str = row['Time'].strftime('%I:%M:%S %p')
            time_str = row['Time']
            # print(time_str)
            date_time_str = f"{date_str} {time_str} NZST"
            value = row[meter]
            date_time_list.append(date_time_str)
            value_list.append(value)
        
        # Create a DataFrame for the current meter and reverse the order of rows
        meter_df = pd.DataFrame({'dates_col': date_time_list, 'values_col': value_list}).iloc[::-1]
        meter_df = meter_df  # Remove the last item
        meter_dict[meter_name] = meter_df
    return meter_dict

def get_relevant_info(time_usage_cols):
    """Returns all the dates and usages for 12AM and 11:55pm"""
    date_val_pair = []
    # Index for adding the first timepoint regardless of its time (Skips the header)
    i = 0
    for row in time_usage_cols[1:]:
        if ('11:55:00 PM' in row[0]) or ('12:00:00 AM' in row[0]) or (i == 0):
            i += 1
            date = row[0][:9] # Always in DD-MMM-YY format
            value = int(row[1][:-4]) # Removes trailing units and converts to usable number
            date_val_pair.append([date, value])

    if len(date_val_pair) % 2 != 0:
        print("WARNING -- Non even entries")

    return date_val_pair

def average(arry):
    return sum(arry)/len(arry)

def plot_water_usage_with_accumulation(data, meter, output_filename, ylims):
    weekday_values = []
    weekend_values = []
    dates = []
    water_usage = []
    for date_str, usage in data:
        dates.append(datetime.strptime(date_str, "%d-%b-%y"))
        water_usage.append(usage)
        if is_weekend_day(date_str):
            weekend_values.append(usage)
        else:
            weekday_values.append(usage)

    # dates = [datetime.strptime(date_str, "%d-%b-%y") for date_str, _ in data]
    # water_usage = [usage for _, usage in data]

    # Create the primary axes for the bar graph
    fig, ax1 = plt.subplots(figsize=(10, 6))

    # Plot the bar graph on the primary axes
    ax1.bar(dates, water_usage, align='center', alpha=0.7, label='Water Usage')
    ax1.set_xlabel('Day')
    ax1.set_ylabel('Water Usage ($m^3$)')  # Using LaTeX notation for m^3
    print(meter)
    # Create the secondary axes for the accumulation line
    ax2 = ax1.twinx()

    # Set the limits for each of the plots based on the maximum sum and reading for that area in the quarter (plus an additional 5% for nicer viewing)
    for key, values in ylims.items():
        if key == meter:
            print(values)
            ax1.set_ylim(0, values[0]+0.05*values[0])
            ax2.set_ylim(0, values[1]+0.05*values[1])

    # if 'Basement - Usage' == meter:
    #     # ax1.set_title('Basement Daily Changes and Cumulative Sum')
    #     ax1.set_ylim(0, 2.5)
    #     ax2.set_ylim(0, 30)
    # elif 'Common Areas - Usage' == meter:
    #     # ax1.set_title('Common Areas Daily Changes and Cumulative Sum')
    #     ax1.set_ylim(0, 50)
    #     ax2.set_ylim(0, 800)
    # elif 'Level 01-08 - Usage' == meter:
    #     # ax1.set_title('Level 01 - 08 Daily Changes and Cumulative Sum')
    #     ax1.set_ylim(0, 55)
    #     ax2.set_ylim(0, 900)
    # elif 'Level 09-18 - Usage' == meter:
    #     # ax1.set_title('Level 09 - 18 Daily Changes and Cumulative Sum')
    #     ax1.set_ylim(0, 55)
    #     ax2.set_ylim(0, 900)
        
    ax1.set_xticks(dates)
    ax1.axhline(average(weekday_values), xmax=0.95, linestyle='--', color='orange', label='Weekday Avg')
    ax1.axhline(average(weekend_values), xmax=0.95, linestyle='--', color='green', label='Weekend Avg')

    # # Plot the weekday averages
    # ax1.plot(dates, [average(weekday_values)]*len(dates), color='orange', linestyle='dashed', label='Weekday Avg')
    # ax1.plot(dates, [average(weekend_values)]*len(dates), color='green', linestyle='dashed', label='Weekend Avg')

    # ax1.set_xticklabels([date.strftime("%d-%b-%y") for date in dates], rotation=45, ha='right')
    ax1.set_xticklabels([i for i in range(1, len(dates)+1)])

    # Accumulate the water usage data
    accumulated_usage = [sum(water_usage[:i + 1]) for i in range(len(water_usage))]

    # Plot the accumulation line on the secondary axes
    ax2.plot(dates, accumulated_usage, color='red', label='Accumulation')
    ax2.set_ylabel('Cumulative Sum ($m^3$)')


    # Combine the legend for both axes
    lines, labels = ax1.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    ax2.legend(lines + lines2, labels + labels2, loc='upper left')

    plt.tight_layout()
    plt.savefig(output_filename)
    plt.close()

def trim_data_dictionary(dictionary, month):
    """ Takes a dictionary containing the meters and all the timepoints (only one per day) for them and returns a similar dictionary the dates and the usage for that date"""
    new_dictionary = {}
    # print(month)
    for meter, data_set in dictionary.items():
        wanted_dates_values = []    
        # Get the dates and values 
        all_dates = data_set['dates_col'].tolist()
        # print(all_dates)
        all_values = data_set['values_col'].tolist()

        for i in range(0, len(all_dates)):
            if 'p.m.' in all_dates[i]:
                if (month in all_dates[i]):
                    wanted_date = all_dates[i][0:9]
                    wanted_value = all_values[i] - all_values[i-1]
                    wanted_dates_values.append([wanted_date, wanted_value])
            else:
                if (month in all_dates[i-1]) and i > 1:
                    wanted_date = all_dates[i-1][0:9]
                    wanted_value = all_values[i] - all_values[i-1]
                    wanted_dates_values.append([wanted_date, wanted_value])
            

        # print(len(wanted_dates_values))



        # dates = [date.replace(' NZST', '') for date in data_set['dates_col'].tolist()]
        # values = [value for value in data_set['values_col'].tolist()]

        # # Limit the dates to be just the month requested
        # wanted_dates = []
        # wanted_values = []
        # for i in range(0, len(dates)):
        #     if month in dates[i]:
        #         wanted_values.append(values[i])
        #         if 'p.m.' in dates[i]:
        #             wanted_dates.append(dates[i].replace('p.m.', 'PM'))
        #         elif 'a.m.' in dates[i]:
        #             wanted_dates.append(dates[i].replace('a.m.', 'AM'))

        # # dates = dates[i:]
        # # values = values[i:]

        
        
        # # The first date and value are set as the past date/value
        # past_date = wanted_dates[0][:9]
        # past_value = wanted_values[0]

        # for i in range(1, len(wanted_dates)):
        #     current_date = wanted_dates[i][:9]
        #     current_value = wanted_values[i]
        #     wanted_dates_values.append([past_date, current_value-past_value])
        #     past_date = current_date
        #     past_value = current_value

        # # Add the last entry
        # timestamp_format = "%d-%b-%y %I:%M:%S %p"
        # timestamp_dt = datetime.strptime(wanted_dates[-1], timestamp_format)
        # end_time = timestamp_dt.strftime("%d-%b-%y (%I%p)")
        # last_usage = values[-1]-past_value
        # wanted_dates_values.append([dates[-1][:9], last_usage])
        # print(f"{meter}\n{wanted_dates_values}\n\n")
        new_dictionary[meter] = wanted_dates_values

    return new_dictionary


def pad_missing_dates(data):
    # Convert the date strings to datetime objects for easy comparison
    dates = [datetime.strptime(row[0], '%d-%b-%y') for row in data[1:-1]]
    datas_dates = [pair[0] for pair in data]
    values = [pair[1] for pair in data]

    # Find the start and end dates in the data
    start_date = min(dates)
    end_date = max(dates)

    # Create a list to store the modified data
    modified_data = [data[0]]  # Header row

    # Loop through the dates and add missing dates with 'NA'
    current_date = start_date
    i = 0
    while current_date <= end_date:
        formatted_date = current_date.strftime('%d-%b-%y')
        if formatted_date not in datas_dates:
            # print(formatted_date, '----', data)
            modified_data.append([formatted_date, 'NA'])
        else:
            modified_data.append([datas_dates[i], values[i]])
            i += 1

        # Move to the next date
        current_date += timedelta(days=1)

    # Add the last row
    modified_data.append(data[-1])

    return modified_data


def set_uniform_spacing(worksheet, start_column, end_column, width):
    for col_idx in range(start_column, end_column + 1):
        column_letter = openpyxl.utils.get_column_letter(col_idx)
        worksheet.column_dimensions[column_letter].width = width


def update_merged_cell_value(worksheet, row, column, value):
    # Get the merged cell range containing the cell at (row, column)
    for cell_range in worksheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row  = cell_range.bounds
        if (min_row <= row <= max_row) and (min_col <= column <= max_col):
            # Unmerge the cells within the range
            worksheet.unmerge_cells(str(cell_range))
            # Update the value in each cell within the merged cell range
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    worksheet.cell(row=r, column=c, value=value)
            # Merge the cells again
            worksheet.merge_cells(start_row=min_row, start_column=min_col,
                                  end_row=max_row, end_column=max_col)
            # Break the loop after updating the merged cell range
            break


def is_weekend_day(date_string):
    # Convert the date string to a date object
    date_obj = datetime.strptime(date_string, "%d-%b-%y").date()

    # Check if the day of the week is Saturday (5) or Sunday (6)
    return date_obj.weekday() in [5, 6]


def write_data_to_excel(dictionary, file_path, month, sheet_name):
    # print(dictionary)
    wb = openpyxl.load_workbook(file_path)
    
    # Check if a sheet with the same name already exists
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])  # Remove the existing sheet
    
    if 'Group' in sheet_name:
        template_sheet = wb["groupingTemplate"]  # Assuming the data should be written to Sheet1
    else:
        template_sheet = wb["template"]  # Assuming the data should be written to Sheet1

    new_sheet_name = sheet_name
    sheet = wb.copy_worksheet(template_sheet)
    sheet.title = new_sheet_name

    update_merged_cell_value(sheet, 1, 2, month)

    # Get the number of days in the given month
    # month_num_days = calendar.monthrange(2023, list(calendar.month_abbr).index(month[:3]))[1]
    month_num_days = len(dictionary[next(iter(dictionary))])

    # Delete excess columns in the template sheet
    heading_cols = 1 # 1 = A, 2 = B, 3 = C
    month_col_start = heading_cols + 1
    last_month_col = 31 + heading_cols # Minus the total column
    if (last_month_col - heading_cols) > month_num_days:
        delete_cols = sheet.iter_cols(min_col=month_col_start + month_num_days, max_col=last_month_col)
        for col in delete_cols:
            last_month_col -= 1
            sheet.delete_cols(col[0].column)

    wb.save(file_path)

    col_end = month_num_days + month_col_start

    # Get the dates in the first row starting from column B and format them in the same format as dictionary dates
    dates = [sheet.cell(row=2, column=col_idx).value for col_idx in range(month_col_start, col_end)]
    print(dates)
    month_year = next(iter(dictionary.values()))[0][0][3:]
    dates = [f'{str(date).zfill(2)}-{month_year}' for date in dates]

    # Create a dictionary to map between the two date formats
    date_map = {}
    for idx, date in enumerate(dates):
        num_date = date
        date_map[num_date] = idx + 2  # Add 2 to match the column index

    print(date_map)

    weekend_fill = PatternFill(start_color='D3DEF1', end_color='D3DEF1', fill_type='solid')

    # Find the column indices of the weekend dates
    weekend_cols = [col_idx for col_idx, date in enumerate(dates, start=month_col_start) if is_weekend_day(date)]

    # Write data to the table
    row_idx = 3
    for meter_code, date_value_pair_list in dictionary.items():
        # Write the meter name in the first column
        meter_name = get_actual_name(meter_code)
        sheet.cell(row=row_idx, column=1, value=meter_name)

        # Write the values corresponding to the dates in the table
        for date_str, value in date_value_pair_list:
            col_idx = date_map[date_str]
            # Remove any 0's
            if value == 0:
                value = ''
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)

        # Apply the weekend_fill to the entire row for the weekend dates
        for col in weekend_cols:
            weekend_cell = sheet.cell(row=row_idx, column=col)
            weekend_cell.fill = weekend_fill

        row_idx += 1

    set_uniform_spacing(sheet, month_col_start, col_end + 2, 4)

    wb.save(file_path)

def write_all_data_to_new_template_excel(dictionary, subgroup_dict_ID, file_path, month, sheet_name, row_start, col_start):
    # print(dictionary)
    wb = openpyxl.load_workbook(file_path)
    
    # Check if a sheet with the same name already exists
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])  # Remove the existing sheet
    
    if 'Group' in sheet_name:
        template_sheet = wb["Group - Template"]  # Assuming the data should be written to Sheet1
    else:
        template_sheet = wb["All - Template"]  # Assuming the data should be written to Sheet1

    new_sheet_name = sheet_name
    sheet = wb.copy_worksheet(template_sheet)
    sheet.title = new_sheet_name

    update_merged_cell_value(sheet, row_start-2, col_start, month)

    # Get the number of days in the given month and the index of these
    month_num_days = len(dictionary[next(iter(dictionary))])
    month_col_start = col_start # 1 = A, 2 = B, 3 = C
    col_end = month_num_days + month_col_start

    dates = [date for date, value in dictionary[next(iter(dictionary))]]
    # Find the column indices of the weekend dates
    weekend_cols = [col_idx for col_idx, date in enumerate(dates, start=month_col_start) if is_weekend_day(date)]
    weekend_fill = PatternFill(start_color='D3DEF1', end_color='D3DEF1', fill_type='solid')

    # Write data to the table
    row_idx = row_start
    col_idx = col_start

    for groupName, add_sub_lists in subgroup_dict_ID.items():
        if 'All' in groupName:
            metersInGroup = add_sub_lists[0]
            update_merged_cell_value(sheet, row_idx, 1, groupName)
            row_idx += 1
            for meterCode in metersInGroup:
                meterData = dictionary[meterCode]
                meterName = get_actual_name(meterCode)
                sheet.cell(row=row_idx, column=col_start-1, value=meterName)
                for date, value in meterData:
                    # if value == 0:
                    #     value = ''
                    cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                    if col_idx in weekend_cols:
                        cell.fill = weekend_fill
                    col_idx += 1
                row_idx += 1
                col_idx = col_start

    set_uniform_spacing(sheet, month_col_start, col_end + 2, 4)

    wb.save(file_path)

# def get_row_idx(sheet, areaName, subtypeName):
    # for index, row in enumerate(sheet.iter_rows(min_row=0, max_row=sheet.max_row, min_col=0, max_col=sheet.max_col)):
    #     for cell in row:
    #         contents = cell.value
    #         if contents == areaName:
    #             areaFlag = True
    #         if contents == subtypeName and areaFlag:

def write_groups_to_new_template_excel(dictionary, file_path, month, new_sheet_name, row_start, col_start):
    """Function can takes a dictionary as input and the template file path and places the data from the dictionary into 1 of 3 different templates depending on what table is wanted"""

    # Load the excel workbook
    wb = openpyxl.load_workbook(file_path)
    
    # Check if a sheet with the same name already exists (deletes duplicates)
    if new_sheet_name in wb.sheetnames:
        wb.remove(wb[new_sheet_name])  # Remove the existing sheet
    
    # Check what type of template to use
    if 'Group' in new_sheet_name:
        template_sheet = wb["Group - Template"]  # Assuming the data should be written to Sheet1
    elif 'Net' in new_sheet_name:
        template_sheet = wb["Net Grouped - Template"]
    else:
        template_sheet = wb["All - Template"]  # Assuming the data should be written to Sheet1

    # Make a copy of the template sheet to edit
    sheet = wb.copy_worksheet(template_sheet)
    sheet.title = new_sheet_name

    # Update the Month Title in sheet
    update_merged_cell_value(sheet, row_start-3, col_start, month)

    # Find out which of the dates are weekends, and make a fill color to be used
    dates = [date for date, value in dictionary[next(iter(dictionary))]]
    weekend_cols = [col_idx for col_idx, date in enumerate(dates, start=col_start) if is_weekend_day(date)]
    weekend_fill = PatternFill(start_color='D3DEF1', end_color='D3DEF1', fill_type='solid')

    # List to store the meter name and maximum reading for it
    name_max_list = {}

    # Write data to the table
    col_idx = col_start
    for fullName, groupData in dictionary.items():
        # Check the names of Groups/Meters and format such that it will find a match in the excel headings column
        splitName = fullName.split(' _ ')
        if len(splitName) == 2:
            groupName = splitName[0]
            subtypeName = splitName[1]
            areaF = False
        else: 
            groupName = 'Not Applicable'
            subtypeName = splitName[0] + ' -'
            areaF = True

        # Place the data for each row in the correct row
        for index, row in enumerate(sheet.iter_rows(min_row=0, max_row=sheet.max_row)):
            row_content = [cell.value for cell in row]
            if row_content[0] == groupName:
                areaF = True
            elif areaF and (subtypeName in str(row_content[0])):
                for date, value in groupData:
                    cell = sheet.cell(row=index+1, column=col_idx, value=value)
                    if col_idx in weekend_cols:
                        cell.fill = weekend_fill
                    col_idx += 1
                col_idx = col_start
                areaF = False

        max_value = max([value for date, value in groupData])
        cuml_sum = sum([value for date, value in groupData])
        name_max_list[fullName] = [max_value, cuml_sum]

    wb.save(file_path)
    return name_max_list

# def write_groups_to_new_template_excel(dictionary, file_path, month, sheet_name, row_start, col_start):
#     # print(dictionary)
#     wb = openpyxl.load_workbook(file_path)
    
#     # Check if a sheet with the same name already exists
#     if sheet_name in wb.sheetnames:
#         wb.remove(wb[sheet_name])  # Remove the existing sheet
    
#     if 'Group' in sheet_name:
#         template_sheet = wb["Group - Template"]  # Assuming the data should be written to Sheet1
#     else:
#         template_sheet = wb["All - Template"]  # Assuming the data should be written to Sheet1

#     new_sheet_name = sheet_name
#     sheet = wb.copy_worksheet(template_sheet)
#     sheet.title = new_sheet_name

#     update_merged_cell_value(sheet, row_start-3, col_start, month)

#     # Get the number of days in the given month and the index of these
#     month_num_days = len(dictionary[next(iter(dictionary))])
#     month_col_start = col_start # 1 = A, 2 = B, 3 = C
#     col_end = month_num_days + month_col_start

#     dates = [date for date, value in dictionary[next(iter(dictionary))]]
#     # Find the column indices of the weekend dates
#     weekend_cols = [col_idx for col_idx, date in enumerate(dates, start=month_col_start) if is_weekend_day(date)]
#     weekend_fill = PatternFill(start_color='D3DEF1', end_color='D3DEF1', fill_type='solid')

#     # Write data to the table
#     row_idx = row_start
#     col_idx = col_start
    
#     for groupName, groupData in dictionary.items():
#         splitName = groupName.split(' - ')
#         # print(splitName)
#         areaName = splitName[0]
#         subtypeName = splitName[1]
#         if 'SUB' not in groupName:
#             update_merged_cell_value(sheet, row_idx, 1, subtypeName)
#             print(row_idx, col_idx)
#             for date, value in groupData:
#                 cell = sheet.cell(row=row_idx, column=col_idx, value=value)
#                 if col_idx in weekend_cols:
#                     cell.fill = weekend_fill
#                 col_idx += 1
#             row_idx += 1
#             col_idx = col_start
#         else:
#             update_merged_cell_value(sheet, row_idx, 1, areaName)
#             row_idx += 1

#     set_uniform_spacing(sheet, month_col_start, col_end + 2, 4)

#     wb.save(file_path)

def plot_dictionary_w_table(dictionary, output_location, ylims):
    """Plots every meter in the dictionary and creates and excel file to store the data for each"""
    for meterName, date_usage in dictionary.items():

        # meterName = get_actual_name(meterName)

        # Set up paths for all the excel files
        output_xlsx_path = f'{output_location}/Excel_Data'
        if not os.path.exists(output_xlsx_path):
            os.makedirs(output_xlsx_path)
            print(f"Folder '{output_xlsx_path}' created.")
        # else:
        #     print(f"Folder '{output_xlsx_path}' already exists.")
        output_xlsx_filename = f'{output_xlsx_path}/{meterName}.xlsx'
        print(output_xlsx_filename)

        # Set up paths for all the png plots
        output_plot_path = f'{output_location}/Plot_Data'
        if not os.path.exists(output_plot_path):
            os.makedirs(output_plot_path)
            print(f"Folder '{output_plot_path}' created.")
        # else:
        #      print(f"Folder '{output_xlsx_path}' already exists.")
        plot_output_name = f'{output_plot_path}/{meterName}.png'
        print(plot_output_name)

        plot_water_usage_with_accumulation(date_usage, meterName, plot_output_name, ylims)

        output_df = pd.DataFrame(date_usage)
        transposed_df = output_df.T
        transposed_df.to_excel(output_xlsx_filename, index=False)


def get_actual_name(meter_code):
    name_map_dict = {
    'BNZ Floors': 'M1',
    'Deloitte':'M2',
    'Cooling Towers':'M3',
    'BNZ retail':'M4',
    'Altezano Café':'M5',
    'Lacoste':'M6',
    'Ben Sherman':'M7',
    'Rockport':'M8',
    'North Face':'M9',
    'Loading dock':'M10',
    'BNZ Showers':'M11',
    'Deloitte Showers':'M12',
    'Harvest BNZ Showers':'M13',
    'Harvest Deloitte Showers': 'M14',
    'Harvest BNZ Retail': 'M15',
    'Harvest B1 & BNZ': 'M16',
    'Harvest - Ground Flr': 'M17',
    'Harvest Deloitte': 'M18',
    'Harvest not used': 'M19',
    'Harvest Domestic top up': 'M20'}
    # Map the meterName from the input data to the actual name if none found then keep code
    meterName = meter_code
    for actual_name, code in name_map_dict.items():
        if meter_code == code:
            meterName = actual_name + f"({meterName})"
    return meterName


def group_meters(original_dict, groups_dict):
    """
    Group meters based on specified combinations, add and subtract values accordingly.

    Parameters:
        original_dict (dict): Original dictionary containing meter names as keys and date_value pairs as values.
        groups_dict (dict): A dictionary where the keys are the names of the subgroups, and the values are lists
                            containing two lists: the first list contains meters whose values should be added together,
                            and the second list contains meters whose values should be subtracted.

    Returns:
        dict: A new dictionary where each key is a subgroup name, and the corresponding value is a list of
              aggregated date_value pairs for all the meters in that subgroup, considering the specified additions
              and subtractions.
    """
    def aggregate_values(meter_list):
        aggregated_data = {}
        for meter in meter_list:
            if meter in original_dict:
                for date, value in original_dict[meter]:
                    if date in aggregated_data:
                        aggregated_data[date] += value
                    else:
                        aggregated_data[date] = value
        return [(date, value) for date, value in aggregated_data.items()]

    grouped_dict = {}

    for group_name, group_lists in groups_dict.items():
        add_meters = group_lists[0]
        subtract_meters = group_lists[1]

        grouped_data = aggregate_values(add_meters)

        if subtract_meters:
            subtracted_data = aggregate_values(subtract_meters)
            for date, value in subtracted_data:
                for idx, (d, v) in enumerate(grouped_data):
                    if d == date:
                        grouped_data[idx] = (d, v - value)
                        break
                else:
                    print("SOMETHING FIHSY")
                    grouped_data.append((date, -value))

        grouped_dict[group_name] = grouped_data

    return grouped_dict

def get_month_list_from_quarter(quarter):
    """Gets a list of the months in a given quarter"""
    if ('Dec' in quarter) or ('Jan' in quarter):
        month_list = ['December', 'January', 'February']
    elif ('Mar' in quarter) or ('May' in quarter):
        month_list = ['March', 'April', 'May']
    elif ('Jun' in quarter) or ('Aug' in quarter):
        month_list = ['June', 'July', 'August']
    elif ('Sep' in quarter) or ('Nov' in quarter):
        month_list = ['September', 'October', 'November']
    return month_list

def get_singular_month(month_input, months_list):
    """Takes the input from the user to select what months to include in the process"""
    # convert to lowercase to make more robust
    month_input = month_input.lower()
    if 'jan' in month_input:
        months_list = ['January']
    elif 'feb' in month_input:
        months_list = ['February']
    elif 'mar' in month_input:
        months_list = ['March']
    elif 'apr' in month_input:
        months_list = ['April']
    elif 'may' in month_input:
        months_list = ['May']
    elif 'jun' in month_input:
        months_list = ['June']
    elif 'jul' in month_input:
        months_list = ['July']
    elif 'aug' in month_input:
        months_list = ['August']
    elif 'sep' in month_input:
        months_list = ['September']
    elif 'oct' in month_input:
        months_list = ['October']
    elif 'nov' in month_input:
        months_list = ['November']
    elif 'dec' in month_input:
        months_list = ['December']
    else:
        months_list = get_month_list_from_quarter(QUARTER)
    return months_list

subgroup_add_sub_dict = {}

subgroup_add_sub_dict['Basement _ Usage'] = [['M11', 'M12'], []]
subgroup_add_sub_dict['Basement _ Harvest'] = [['M13', 'M14'], []]
subgroup_add_sub_dict['Basement _ Net'] = [['M11', 'M12'], ['M13', 'M14']]

subgroup_add_sub_dict['Common Areas _ Usage'] = [['M5', 'M6', 'M7', 'M8', 'M9', 'M4', 'M3', 'M10'], []]
subgroup_add_sub_dict['Common Areas _ Harvest'] = [['M17', 'M15', 'M19', 'M20'], []]
subgroup_add_sub_dict['Common Areas _ Net'] = [['M5', 'M6', 'M7', 'M8', 'M9', 'M4', 'M3', 'M10'],  ['M17', 'M15', 'M19', 'M20']]

subgroup_add_sub_dict['Levels 01 - 08 _ Usage'] = [['M1'], []]
subgroup_add_sub_dict['Levels 01 - 08 _ Harvest'] = [['M16'], []]
subgroup_add_sub_dict['Levels 01 - 08 _ Net'] = [['M1'], ['M16']]

subgroup_add_sub_dict['Levels 09 - 18 _ Usage'] = [['M2', 'M2-09-1', 'M2-09-2', 'M2-10-1', 'M2-10-2'], []]
subgroup_add_sub_dict['Levels 09 - 18 _ Harvest'] = [['M18', 'M18-09', 'M18-10'], []]
subgroup_add_sub_dict['Levels 09 - 18 _ Net'] = [['M2', 'M2-09-1', 'M2-09-2', 'M2-10-1', 'M2-10-2'], ['M18', 'M18-09', 'M18-10']]

### SHORTCUT ###
QUARTER = 'Dec-Feb2024'
data_frame = pd.read_excel(f'WATER/{QUARTER}/WaterMetersDec-Feb.xlsx', 'Data')
###

#### Inputs for running function ###
# while True:
#     try:
#         # add some input here to ask the user to select the folder and file to read the data from
#         QUARTER = input("Enter the name of the Folder for the Quarter (eg. 'Dec-Feb2024'): ")
#         filename_guess = f"WaterMeters{QUARTER}.xlsx"
#         y_n = input(f"Is the name of the input datafile '{filename_guess}' (Y/N): ")
#         if ('Y' in y_n) or ('y' in y_n):
#             input_file = filename_guess
#         else:
#             input_file = input(f"Enter the name of the input file (including .xlsx): ")a
#         data_frame = pd.read_excel(f'WATER/{QUARTER}/{input_file}', 'Data')
#         break  # Exit the loop if file is successfully opened
#     except FileNotFoundError:
#         print("File not found. Please enter a valid filename.")
#     except Exception as e:
#         print("An error occurred:", e)

# Get the data for 1 month at a time and generate the outputs for each 


functions_toDo = input("What outputs do you want (a = All, g = groups table, e = each meter table, n = net table, p = plots): ")
month_input = input("Press 'Enter' to process whole Quarter, otherwise type the month you would like to process: ")

months_list = get_singular_month(month_input, QUARTER)
input(f"Beginning processing {months_list} ...  (press 'Enter' to start)")
#### ^^^^ ###


# Load template excel workbook and create a new workbook to store all the results for the quarter in (Create a Output folder if needed)
template_xlsx = f'WATER/water_meter_templates.xlsx'
wb = openpyxl.load_workbook(template_xlsx)
output_all_table_name = 'water_meter_output_table.xlsx'
output_folder = f'WATER/{QUARTER}/Outputs'
if not os.path.exists(output_folder):
        os.makedirs(output_folder)
        print(f"Folder '{output_folder}' created.")

water_tables_xlsx = f'{output_folder}/{output_all_table_name}'
wb.save(water_tables_xlsx)


# Extract all the data from the input file and allocate into a dictionary

meters_data_dict = extract_meter_values(data_frame)
# print(meters_data_dict)

# List to store all the lists of quarterly maximums
quarterly_maxs = []

for desired_month in months_list:
    # desired_month = 'February'Dec
    # # desired_month = 'January'Dec-F
    # # desired_month = 'December'
    meters_data_dict_to_plot = trim_data_dictionary(meters_data_dict, desired_month[:3])

    sub_group_dict = group_meters(meters_data_dict_to_plot, subgroup_add_sub_dict)

    """Here are the 3 possible callings of the function that writes all the data into various spreadsheets"""
    if ('a' in functions_toDo) or ('g' in functions_toDo):
        # Adds the grouped Usage and Harvest data to the Groups Table (MAIN TABLE USED IN REPORT)
        quarterly_maxs.append(write_groups_to_new_template_excel(sub_group_dict, water_tables_xlsx, desired_month, "Groups - "+desired_month[:3], row_start=5, col_start=3))
    if ('a' in functions_toDo) or ('n' in functions_toDo):
        # Adds the grouped Usage, Harvest and Net data to the Net Table.
        quarterly_maxs.append(write_groups_to_new_template_excel(sub_group_dict, water_tables_xlsx, desired_month, "Nets - "+desired_month[:3], row_start=5, col_start=3))
    if ('a' in functions_toDo) or ('e' in functions_toDo):
        # Adds every meter to the All table, plots every meter individually and stores data in excel table
        quarterly_maxs.append(write_groups_to_new_template_excel(meters_data_dict_to_plot, water_tables_xlsx, desired_month, "All - "+desired_month[:3], row_start=5, col_start=3))


# Find all the ylims from the maximum readings from each of the meters
ylims_dict = {}
# print(quarterly_maxs)
for g_dict in quarterly_maxs:
    for key, values in g_dict.items():
        try:
            ylims_dict[key]
        except KeyError:
            ylims_dict[key] = [0, 0]
        if values[0] > ylims_dict[key][0]:
            ylims_dict[key][0] = values[0]
        if values[1] > ylims_dict[key][1]:
            ylims_dict[key][1] = values[1]


print(ylims_dict)

# Plot for each month considering the ylims
for desired_month in months_list:
    meters_data_dict_to_plot = trim_data_dictionary(meters_data_dict, desired_month[:3])

    sub_group_dict = group_meters(meters_data_dict_to_plot, subgroup_add_sub_dict)
    
    ### Plots every meter individually
    if ('a' in functions_toDo) or (('e' in functions_toDo) and ('p' in functions_toDo)):
        plot_dictionary_w_table(dictionary=meters_data_dict_to_plot, output_location=f'WATER/{QUARTER}/Outputs/{desired_month}_Each_Meter', ylims=ylims_dict)

    ### Plots just the groups of meters
    if ('a' in functions_toDo) or ('p' in functions_toDo):
        #  Plots every group and stores the data in excel table
        plot_dictionary_w_table(dictionary=sub_group_dict, output_location=f'WATER/{QUARTER}/Outputs/{desired_month}_Groups', ylims=ylims_dict)
