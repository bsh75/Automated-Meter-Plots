# Step 1: Read the CSV file and extract columns A and D
import csv
from openpyxl import Workbook
import matplotlib.pyplot as plt
from datetime import datetime
import os


def extract_time_and_value(filename, header_index):
    columns_A_and_D = []

    with open(filename, 'r') as csvfile:
        csv_reader = csv.reader(csvfile)
        for row in csv_reader:
            if len(row) >= header_index:
                columns_A_and_D.append([row[0], row[3]])

    return columns_A_and_D

def get_relevant_info(csv_file, header_row):
    """Removes all the unnecessary times and """
    time_value = extract_time_and_value(csv_file, header_index=header_row)
    date_val_pair = []
    # Index for adding the first timepoint regardless of its time (Skips the header)
    i = 0
    for row in time_value[1:]:
        if ('11:55:00 PM' in row[0]) or ('12:00:00 AM' in row[0]) or (i == 0):
            i += 1
            date = row[0][:9] # Always in DD-MMM-YY format
            value = int(row[1][:-4]) # Removes trailing units and converts to usable number
            date_val_pair.append([date, value])

    if len(date_val_pair) % 2 != 0:
        print("WARNING -- Non even entries")

    return date_val_pair

def modded_csv_to_xlsx_data_return(csv_filename, xlsx_filename):
    # Read CSV data and convert the second column to integers
    csv_data = []
    with open(csv_filename, 'r', encoding='utf-8-sig') as csv_file:  # Use utf-8-sig to remove BOM
        csv_reader = csv.reader(csv_file)
        first_row = next(csv_reader)  # Read and store the first row
        for row in csv_reader:
            # Convert the second column to an integer
            row[1] = int(row[1])
            csv_data.append(row)

    # Create a new Excel workbook
    wb = Workbook()
    ws = wb.active

    # Write CSV data (excluding the first row) to the worksheet
    for row_idx, row_data in enumerate(csv_data):
        for col_idx, cell_value in enumerate(row_data):
            ws.cell(row=row_idx+2, column=col_idx+1, value=cell_value)  # Start from row 2

    # Write the first row to the worksheet
    for col_idx, cell_value in enumerate(first_row):
        ws.cell(row=1, column=col_idx+1, value=cell_value)

    # Save the workbook as an Excel file
    wb.save(xlsx_filename)

    # Return the data excluding the first row
    return csv_data

def plot_water_usage_with_accumulation(data, meter, output_filename):
    dates = [datetime.strptime(date_str, "%d-%b-%y") for date_str, _ in data]
    water_usage = [usage for _, usage in data]

    # Create the primary axes for the bar graph
    fig, ax1 = plt.subplots(figsize=(10, 6))

    # Plot the bar graph on the primary axes
    ax1.bar(dates, water_usage, align='center', alpha=0.7, label='Water Usage')
    ax1.set_xlabel('Date')
    ax1.set_ylabel('Water Usage ($m^3$)')  # Using LaTeX notation for m^3
    ax1.set_title(f'{meter} Water Usage')
    ax1.set_xticks(dates)
    ax1.set_xticklabels([date.strftime("%d-%b-%y") for date in dates], rotation=45, ha='right')

    # Create the secondary axes for the accumulation line
    ax2 = ax1.twinx()

    # Accumulate the water usage data
    accumulated_usage = [sum(water_usage[:i + 1]) for i in range(len(water_usage))]

    # Plot the accumulation line on the secondary axes
    ax2.plot(dates, accumulated_usage, color='red', label='Accumulation')
    ax2.set_ylabel('Accumulation ($m^3$)')
    
    # Combine the legend for both axes
    lines, labels = ax1.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    ax2.legend(lines + lines2, labels + labels2, loc='upper left')

    plt.tight_layout()
    plt.savefig(output_filename)
    plt.close()


start_sequence = '80QWaterUsage'
file_type = '.csv'

already_plot_ready_format = ['M01', 'M02', 'M03']


file_list = os.listdir('80Q - Water Meters')
for file in file_list:
    if file.endswith(file_type) and file.startswith(start_sequence):
        # IO files
        meterName = file.replace(file_type, '').replace(start_sequence, '')
        input_csv_file = f'80Q - Water Meters/80QWaterUsage{meterName}.csv'
        output_xlsx_filename = f'80Q - Water Meters/Plot Data/{meterName}.xlsx'
        plot_output_name = f'80Q - Water Meters/Plot Data/{meterName}.png'

        if meterName in already_plot_ready_format:
            date_usage = modded_csv_to_xlsx_data_return(input_csv_file, output_xlsx_filename)
            print(f'File {file} is already in plot ready format')
        else:
            print(f'File {file} is in raw format')

            # Get the date and water usage for each day
            wb = Workbook()
            sheet = wb.active


            sheet.append(["Date", "Water Usage"])
            date_val = get_relevant_info(input_csv_file, header_row=4)
            num_dates = len(date_val)
            date_usage = []

            for i in range(0, num_dates-1):
                current_pair = date_val[i]
                next_pair = date_val[i+1]
                if current_pair[0] == next_pair[0]:
                    date = current_pair[0]
                    water_usage = next_pair[1] - current_pair[1]
                    pair = [date, water_usage]
                    date_usage.append(pair)
                    sheet.append(pair)

            wb.save(output_xlsx_filename)


        plot_water_usage_with_accumulation(date_usage, meterName, plot_output_name)
