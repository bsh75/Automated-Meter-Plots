# Step 1: Read the CSV file and extract columns A and D
import csv
import pandas as pd
import matplotlib.pyplot as plt
from datetime import datetime, timedelta
import os
from openpyxl import Workbook

def modded_csv_to_xlsx_data_return(csv_filename, month):
    # Read CSV data and convert the second column to integers
    csv_data = []
    with open(csv_filename, 'r', encoding='utf-8-sig') as csv_file:  # Use utf-8-sig to remove BOM
        csv_reader = csv.reader(csv_file)
        first_row = next(csv_reader)  # Read and store the first row
        for row in csv_reader:
            # Convert the second column to an integer
            row[1] = int(row[1])
            if month in row[0]:
                # print(row)
                csv_data.append(row)
    return csv_data
    # # Create a new Excel workbook
    # wb = Workbook()
    # ws = wb.active

    # # Write CSV data (excluding the first row) to the worksheet
    # for row_idx, row_data in enumerate(csv_data):
    #     for col_idx, cell_value in enumerate(row_data):
    #         ws.cell(row=row_idx+2, column=col_idx+1, value=cell_value)  # Start from row 2

    # # Write the first row to the worksheet
    # for col_idx, cell_value in enumerate(first_row):
    #     ws.cell(row=1, column=col_idx+1, value=cell_value)

    # # Save the workbook as an Excel file
    # wb.save(xlsx_file_path)

    # # Return the data excluding the first row
    # return csv_data


def get_meters_data(df, gap):
    """Function which looks at a file to find its name and data"""
    name_start = 'history:BMS_Supervisor/HYD_METER_'
    columns_A_and_D = []
    start_indexs = []
    meter_names = []
    # data_sets = []
    name_data_dict = {}
    for index, row in df.iterrows():
        if str(row[0]).startswith(name_start):
            start_indexs.append(index)
            meter_names.append(df.iloc[index,0].replace(name_start, ''))

    # Now go through each of the found meters and add the data to the dictionary for each one
    if len(start_indexs) > 1:
        for i in range(0, len(start_indexs)-1):
            name_data_dict[meter_names[i]] = df.iloc[start_indexs[i]+gap:start_indexs[i+1], [0, 3]]
    data_to_add = df.iloc[start_indexs[-1]+gap:, [0, 3]]
    name_data_dict[meter_names[-1]] = data_to_add
    return name_data_dict

    # with open(filename, 'r') as csvfile:
    #     csv_reader = csv.reader(csvfile)
    #     for row in csv_reader:
    #         if len(row) >= header_index:
    #             columns_A_and_D.append([row[0], row[3]])

    # return columns_A_and_D

def get_relevant_info(time_usage_cols):
    """Returns all the dates and usages for 12AM and 11:55pm"""
    date_val_pair = []
    # Index for adding the first timepoint regardless of its time (Skips the header)
    i = 0
    for row in time_usage_cols[1:]:
        if ('11:55:00 PM' in row[0]) or ('12:00:00 AM' in row[0]) or (i == 0):
            i += 1
            date = row[0][:9] # Always in DD-MMM-YY format
            value = int(row[1][:-4]) # Removes trailing units and converts to usable number
            date_val_pair.append([date, value])

    if len(date_val_pair) % 2 != 0:
        print("WARNING -- Non even entries")

    return date_val_pair

def plot_water_usage_with_accumulation(data, meter, output_filename):
    dates = [datetime.strptime(date_str, "%d-%b-%y") for date_str, _ in data]
    water_usage = [usage for _, usage in data]

    # Create the primary axes for the bar graph
    fig, ax1 = plt.subplots(figsize=(10, 6))

    # Plot the bar graph on the primary axes
    ax1.bar(dates, water_usage, align='center', alpha=0.7, label='Water Usage')
    ax1.set_xlabel('Date')
    ax1.set_ylabel('Water Usage ($m^3$)')  # Using LaTeX notation for m^3
    ax1.set_title(f'{meter} Water Usage')
    ax1.set_xticks(dates)
    ax1.set_xticklabels([date.strftime("%d-%b-%y") for date in dates], rotation=45, ha='right')

    # Create the secondary axes for the accumulation line
    ax2 = ax1.twinx()

    # Accumulate the water usage data
    accumulated_usage = [sum(water_usage[:i + 1]) for i in range(len(water_usage))]

    # Plot the accumulation line on the secondary axes
    ax2.plot(dates, accumulated_usage, color='red', label='Accumulation')
    ax2.set_ylabel('Accumulation ($m^3$)')
    
    # Combine the legend for both axes
    lines, labels = ax1.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    ax2.legend(lines + lines2, labels + labels2, loc='upper left')

    plt.tight_layout()
    plt.savefig(output_filename)
    plt.close()

def trim_data_dictionary(dictionary, month):
    """ Takes a dictionary containing the meters and all the timepoints for them and returns a similar dictionary 
        containing only the useful timepoints (12am 11:55pm and the first and last)"""
    new_dictionary = {}
    for meter, data_set in dictionary.items():
        # Get the dates and values 
        dates = [date.replace(' NZST', '') for date in data_set['dates_col'].tolist()]
        print(data_set['values_col'])
        values = [int(value[0:-3]) for value in data_set['values_col'].tolist()]

        # Limit the dates to be just the month requested
        for i in range(0, len(dates)):
            if month in dates[i]:
                break
        dates = dates[i:]
        values = values[i:]

        wanted_dates_values = []
        
        # The first date and value are set as the past date/value
        past_date = dates[0][:9]
        past_value = values[0]

        for i in range(1, len(dates)):
            current_date = dates[i][:9]
            current_value = values[i]
            if (current_date != past_date):
                # If current date is different to the past the take the difference in values to be the past dates usage
                wanted_dates_values.append([past_date, current_value-past_value])
                past_date = current_date
                past_value = current_value
            else:
                continue

        # Add the last entry
        timestamp_format = "%d-%b-%y %I:%M:%S %p"
        timestamp_dt = datetime.strptime(dates[-1], timestamp_format)
        end_time = timestamp_dt.strftime("%d-%b-%y (%I%p)")
        last_usage = values[-1]-past_value
        wanted_dates_values.append([dates[-1][:9], last_usage])
        new_dictionary[meter] = wanted_dates_values

    return new_dictionary, end_time


def pad_missing_dates(data):
    # Convert the date strings to datetime objects for easy comparison
    dates = [datetime.strptime(row[0], '%d-%b-%y') for row in data[1:-1]]
    datas_dates = [pair[0] for pair in data]
    values = [pair[1] for pair in data]

    # Find the start and end dates in the data
    start_date = min(dates)
    end_date = max(dates)

    # Create a list to store the modified data
    modified_data = [data[0]]  # Header row

    # Loop through the dates and add missing dates with 'NA'
    current_date = start_date
    i = 0
    while current_date <= end_date:
        formatted_date = current_date.strftime('%d-%b-%y')
        if formatted_date not in datas_dates:
            # print(formatted_date, '----', data)
            modified_data.append([formatted_date, 'NA'])
        else:
            modified_data.append([datas_dates[i], values[i]])
            i += 1

        # Move to the next date
        current_date += timedelta(days=1)

    # Add the last row
    modified_data.append(data[-1])

    return modified_data


meter_name_map = {
    'BNZ Floors': 'M1',
    'Deloitte':'M2',
    'Cooling Towers':'M3',
    'BNZ retail':'M4',
    'Altezano Café':'M5',
    'Lacoste':'M6',
    'Ben Sherman':'M7',
    'Rockport':'M8',
    'North Face':'M9',
    'Loading dock':'M10',
    'BNZ Showers':'M11',
    'Deloitte Showers':'M12',
    'Harvest BNZ Showers':'M13',
    'Harvest Deloitte Showers': 'M14',
    'Harvest BNZ Retail': 'M15',
    'Harvest B1 & BNZ': 'M16',
    'Harvest - Ground Flr': 'M17',
    'Harvest Deloitte': 'M18',
    'Harvest not used': 'M19',
     'Harvest Domestic top up': 'M20'
}

start_sequence = '80QWaterUsage'
file_type = '.csv'
folder = '80Q - Water Meters'
output_folder = 'Plot Data'
name_gap = 4
file_list = os.listdir(folder)
desired_month = 'Jul'

already_plot_ready_format = [f'{start_sequence}M01{file_type}', f'{start_sequence}M02{file_type}', f'{start_sequence}M03{file_type}']

water_meter_table_data = []
first = True

for file in file_list[10:11]:
    if file.endswith(file_type) and file.startswith(start_sequence):
        input_csv_path = f'{folder}/{file}'

        if file in already_plot_ready_format:
            print()
            # meters_data_dict_to_plot = {}
            # # SPECIAL CASE FOR FILES NATALIE MODIFIED ALREADY (REMOVE IN FUTURE)
            # date_usage = modded_csv_to_xlsx_data_return(input_csv_path, desired_month)
            # meterName = file.replace(start_sequence, '').replace(file_type, '').replace('0', '')
            # meters_data_dict_to_plot[meterName] = date_usage

            # end_time = 'Not applicable'
            # print(f'File {file} is already in plot ready format')
        else:
            # NORMAL BEHAVIOUS
            print(f'File {file} is in raw format')
            data_frame = pd.read_csv(input_csv_path, names=["dates_col", "col2", "col3", "values_col"])
            meters_data_dict = get_meters_data(data_frame, name_gap)
            print("Data: ", type(meters_data_dict['M4']), meters_data_dict)
            meters_data_dict_to_plot, end_time = trim_data_dictionary(meters_data_dict, desired_month)

        for meterName, date_usage in meters_data_dict_to_plot.items():

            # Map the meterName from the input data to the actual name
            for actual_name, code in meter_name_map.items():
                if meterName in code:
                    meterName = actual_name + f"({meterName})"

            timestamp_format = "%d-%b-%y %I:%M:%S %p"

            output_xlsx_filename = f'{folder}/{output_folder}/{meterName}.xlsx'
            plot_output_name = f'{folder}/{output_folder}/{meterName}.png'
            
            print(f'New files created {output_xlsx_filename} and plot {plot_output_name}')
            plot_water_usage_with_accumulation(date_usage, meterName, plot_output_name)
            date_usage.insert(0, ['Date',  'Water Usage (m\u00b3)'])
            date_usage.append(['Up until: ', end_time])
            output_df = pd.DataFrame(date_usage)
            output_df.to_excel(output_xlsx_filename, index=False)
            # print(date_usage)
            date_usage = pad_missing_dates(date_usage)

            # Add data to the water meter table 
            # Get the list of values
            values = [row[1] for row in date_usage[1:-1]]
            values.insert(0, meterName)
            if first:
                print("FIrst file ======= {}".format(file))
                # Get the list of dates
                dates_header_for_table = [row[0] for row in date_usage[:-1]]
                # print(dates_header_for_table)
                first = False
                water_meter_table_data.append(dates_header_for_table)
            water_meter_table_data.append(values)

# print(water_meter_table_data)
combined_df = pd.DataFrame(water_meter_table_data)
combined_df.to_excel('Combined water meter data.xlsx', index=False)

        


        



